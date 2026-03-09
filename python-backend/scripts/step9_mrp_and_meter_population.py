#!/usr/bin/env python3
"""
Step 9 & 10: MRP Data Population + Meter Readings & Plant Performance

Three sequential phases:
  A) Meter Readings CSV → meter_aggregate
  B) MRP Formula OCR + Monthly Data → clause_tariff + reference_price
  C) Plant Performance Enrichment (partial Step 10)

Usage:
  python scripts/step9_mrp_and_meter_population.py [--dry-run] [--no-cache] [--phase A|B|C|all] [--project SAGE_ID]
"""

import argparse
import csv
import hashlib
import io
import json
import logging
import os
import re
import sys
import base64
from calendar import monthrange
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from uuid import uuid4

from dotenv import load_dotenv
load_dotenv()

import anthropic
import openpyxl
from psycopg2.extras import RealDictCursor, execute_values

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
BACKEND_DIR = SCRIPT_DIR.parent
PROJECT_ROOT = BACKEND_DIR.parent

sys.path.insert(0, str(BACKEND_DIR))
from db.database import init_connection_pool, close_connection_pool, get_db_connection

REPORT_DIR = BACKEND_DIR / 'reports' / 'cbe-population'
REPORT_DIR.mkdir(parents=True, exist_ok=True)
IMAGE_DIR = REPORT_DIR / 'step9_mrp_images'
IMAGE_DIR.mkdir(parents=True, exist_ok=True)
OCR_CACHE_DIR = REPORT_DIR / 'step9_ocr_cache'
OCR_CACHE_DIR.mkdir(parents=True, exist_ok=True)

METER_CSV = PROJECT_ROOT / 'CBE_data_extracts' / 'Data Extracts' / 'FrontierMind Extracts_meter readings.csv'
MRP_XLSX = PROJECT_ROOT / 'CBE_data_extracts' / 'MRP' / 'Sage Contract Extracts market Ref pricing data.xlsx'

ORG_ID = 1

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger('step9')

# ---------------------------------------------------------------------------
# Country → currency mapping (for MRP currency inference)
# ---------------------------------------------------------------------------
COUNTRY_CURRENCY = {
    'Ghana': 'GHS',
    'Kenya': 'KES',
    'Nigeria': 'NGN',
    'Sierra Leone': 'SLE',
    'Madagascar': 'MGA',
    'Egypt': 'EGP',
    'Mozambique': 'MZN',
    'Rwanda': 'RWF',
    'Somalia': 'USD',
    'Zimbabwe': 'USD',
    'DRC': 'CDF',
}

# Sage ID aliases (sheet name → canonical sage_id)
SAGE_ALIASES = {
    'MOH001': 'MOH01',
    'MOH01': 'MOH01',
    'KAS01': 'KAS01',
    'UTK01': 'UTK01',
    'UGL01': 'UGL01',
    'TBM01': 'TBM01',
    'GBL01': 'GBL01',
    'JAB01': 'JAB01',
    'NBL01': 'NBL01',
    'NBL02': 'NBL02',
}

# Projects known to have dual-section MRP (Grid + Generator)
DUAL_SECTION_PROJECTS = {'GBL01', 'JAB01'}
# Generator-only projects (grid MRP = 0)
GENERATOR_ONLY_PROJECTS = {'NBL01', 'NBL02'}

# Projects already populated — compare but don't overwrite
ALREADY_POPULATED = {'KAS01', 'MOH01'}


# =========================================================================
# Phase A: Meter Readings CSV → meter_aggregate
# =========================================================================

def phase_a(dry_run: bool, project_filter: Optional[str] = None) -> Dict[str, Any]:
    """Parse meter readings CSV and insert into meter_aggregate."""
    log.info('=== Phase A: Meter Readings → meter_aggregate ===')
    result = {
        'phase': 'A',
        'rows_parsed': 0,
        'rows_inserted': 0,
        'rows_updated': 0,
        'rows_skipped': 0,
        'unresolved_contract_lines': [],
        'unresolved_billing_periods': [],
        'projects_covered': set(),
        'errors': [],
    }

    if not METER_CSV.exists():
        result['errors'].append(f'Meter CSV not found: {METER_CSV}')
        log.error(result['errors'][-1])
        return result

    # ---- Load lookup dicts from DB ----
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '300s'")

        # external_line_id → contract_line row
        cur.execute("""
            SELECT cl.id, cl.external_line_id, cl.contract_id, cl.energy_category,
                   cl.meter_id, cl.billing_product_id,
                   c.project_id, p.sage_id
            FROM contract_line cl
            JOIN contract c ON c.id = cl.contract_id
            JOIN project p ON p.id = c.project_id
            WHERE cl.organization_id = %s AND cl.external_line_id IS NOT NULL
        """, (ORG_ID,))
        cl_by_ext = {}
        for row in cur.fetchall():
            cl_by_ext[row['external_line_id']] = dict(row)

        # (year, month) → billing_period row
        cur.execute("SELECT id, start_date, end_date FROM billing_period ORDER BY start_date")
        bp_rows = cur.fetchall()
        bp_by_ym: Dict[Tuple[int, int], Dict] = {}
        for bp in bp_rows:
            if bp['start_date']:
                bp_by_ym[(bp['start_date'].year, bp['start_date'].month)] = dict(bp)

        # Existing meter_aggregate for dedup check
        cur.execute("""
            SELECT contract_line_id, billing_period_id
            FROM meter_aggregate
            WHERE organization_id = %s AND period_type = 'monthly'
        """, (ORG_ID,))
        existing_ma = {(r['contract_line_id'], r['billing_period_id']) for r in cur.fetchall()}

    # ---- Parse CSV ----
    rows_to_insert = []
    with open(METER_CSV, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            result['rows_parsed'] += 1
            ext_id = row.get('CONTRACT_LINE_UNIQUE_ID', '').strip()
            bill_date_str = row.get('BILL_DATE', '').strip()

            # Resolve contract_line
            cl = cl_by_ext.get(ext_id)
            if not cl:
                result['unresolved_contract_lines'].append({
                    'row': i + 2,
                    'external_line_id': ext_id,
                    'customer': row.get('CUSTOMER_NUMBER', ''),
                })
                result['rows_skipped'] += 1
                continue

            sage_id = cl['sage_id']
            if project_filter and sage_id != project_filter:
                result['rows_skipped'] += 1
                continue

            # Parse bill date → first-of-month
            try:
                bd = datetime.strptime(bill_date_str, '%Y/%m/%d')
            except ValueError:
                try:
                    bd = datetime.strptime(bill_date_str, '%Y-%m-%d')
                except ValueError:
                    result['errors'].append(f'Row {i+2}: bad date {bill_date_str}')
                    result['rows_skipped'] += 1
                    continue

            bp = bp_by_ym.get((bd.year, bd.month))
            if not bp:
                result['unresolved_billing_periods'].append({
                    'row': i + 2,
                    'date': bill_date_str,
                    'year_month': f'{bd.year}-{bd.month:02d}',
                })
                result['rows_skipped'] += 1
                continue

            # Check for existing row
            if (cl['id'], bp['id']) in existing_ma:
                result['rows_skipped'] += 1
                continue

            # Parse numeric fields
            def _dec(val):
                try:
                    v = str(val).strip().replace(',', '')
                    return Decimal(v) if v else Decimal('0')
                except (InvalidOperation, ValueError):
                    return Decimal('0')

            opening = _dec(row.get('OPENING_READING', 0))
            closing = _dec(row.get('CLOSING_READING', 0))
            utilized = _dec(row.get('UTILIZED_READING', 0))
            discount = _dec(row.get('DISCOUNT_READING', 0))
            sourced = _dec(row.get('SOURCED_ENERGY', 0))
            total_prod = utilized - discount - sourced

            energy_cat = cl['energy_category']  # 'metered' or 'available'
            metered_avail = (row.get('METERED_AVAILABLE', '') or '').strip().lower()

            # Route to correct column based on energy_category
            energy_kwh = total_prod if energy_cat == 'metered' else None
            available_kwh = total_prod if energy_cat == 'available' else None

            period_start = bp['start_date']
            _, last_day = monthrange(period_start.year, period_start.month)
            period_end = date(period_start.year, period_start.month, last_day)

            rec = {
                'billing_period_id': bp['id'],
                'contract_line_id': cl['id'],
                'meter_id': cl.get('meter_id'),
                'period_type': 'monthly',
                'period_start': period_start,
                'period_end': period_end,
                'energy_kwh': energy_kwh,
                'available_energy_kwh': available_kwh,
                'total_production': total_prod,
                'opening_reading': opening,
                'closing_reading': closing,
                'utilized_reading': utilized,
                'discount_reading': discount,
                'sourced_energy': sourced,
                'source_system': 'snowflake',
                'source_metadata': json.dumps({
                    'external_reading_id': row.get('METER_READING_UNIQUE_ID', ''),
                    'product_desc': row.get('PRODUCT_DESC', ''),
                    'customer_number': row.get('CUSTOMER_NUMBER', ''),
                    'facility': row.get('FACILITY', ''),
                    'contract_number': row.get('CONTRACT_NUMBER', ''),
                    'metered_available': metered_avail,
                }),
                'organization_id': ORG_ID,
                'unit': 'kWh',
            }
            rows_to_insert.append(rec)
            result['projects_covered'].add(sage_id)

    log.info(f'Parsed {result["rows_parsed"]} rows, {len(rows_to_insert)} to insert, '
             f'{len(result["unresolved_contract_lines"])} unresolved contract lines, '
             f'{len(result["unresolved_billing_periods"])} unresolved billing periods')

    if dry_run:
        result['rows_inserted'] = len(rows_to_insert)
        result['projects_covered'] = sorted(result['projects_covered'])
        log.info(f'[DRY RUN] Would insert {len(rows_to_insert)} meter_aggregate rows')
        return result

    # ---- Batch insert ----
    if rows_to_insert:
        cols = [
            'billing_period_id', 'contract_line_id', 'meter_id',
            'period_type', 'period_start', 'period_end',
            'energy_kwh', 'available_energy_kwh', 'total_production',
            'opening_reading', 'closing_reading', 'utilized_reading',
            'discount_reading', 'sourced_energy',
            'source_system', 'source_metadata', 'organization_id', 'unit',
        ]
        insert_sql = f"""
            INSERT INTO meter_aggregate ({', '.join(cols)})
            VALUES %s
            ON CONFLICT DO NOTHING
        """
        values = [tuple(r[c] for c in cols) for r in rows_to_insert]

        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute("SET statement_timeout = '300s'")
            execute_values(cur, insert_sql, values, page_size=100)
            result['rows_inserted'] = cur.rowcount
            log.info(f'Inserted {cur.rowcount} meter_aggregate rows')

    result['projects_covered'] = sorted(result['projects_covered'])
    return result


# =========================================================================
# Phase B: MRP Formula OCR + Monthly Data
# =========================================================================

def _extract_images_from_sheet(wb: openpyxl.Workbook, sheet_name: str) -> List[bytes]:
    """Extract embedded images from an openpyxl worksheet."""
    ws = wb[sheet_name]
    images = []
    for img in ws._images:
        # openpyxl image objects have a _data() method or ref attribute
        try:
            img_data = img._data()
        except Exception:
            try:
                img_data = img.ref.getvalue() if hasattr(img.ref, 'getvalue') else img.ref.read()
            except Exception as e:
                log.warning(f'Could not extract image from {sheet_name}: {e}')
                continue
        if img_data:
            images.append(img_data)
    return images


def _ocr_image_claude(image_bytes: bytes, sage_id: str, use_cache: bool = True) -> Dict[str, Any]:
    """OCR a formula screenshot via Claude vision, with disk cache."""
    img_hash = hashlib.sha256(image_bytes).hexdigest()
    cache_path = OCR_CACHE_DIR / f'{img_hash}.json'

    if use_cache and cache_path.exists():
        log.info(f'  OCR cache hit for {sage_id} ({img_hash[:12]})')
        return json.loads(cache_path.read_text())

    log.info(f'  Sending image to Claude vision for {sage_id} ({len(image_bytes)} bytes)')

    client = anthropic.Anthropic()
    b64 = base64.b64encode(image_bytes).decode('utf-8')

    prompt = """Analyze this MRP (Market Reference Price) formula screenshot from a solar PPA contract.

Extract the following structured information:

1. **mrp_method**: One of:
   - "utility_variable_charges_tou" (sum of variable energy charges only, excluding demand/fixed/VAT)
   - "utility_total_charges" (sum of all non-tax charges)
   - "generator_cost" (diesel generator or backup power cost)
   - "blended_grid_generator" (combination of grid + generator)

2. **mrp_included_components**: List of tariff components INCLUDED (e.g., "energy_charge", "fuel_surcharge", "lighting_levy", "subsidy")

3. **mrp_excluded_components**: List of components EXCLUDED (e.g., "VAT", "demand_charge", "fixed_charge", "NHIL", "GETFUND")

4. **mrp_exclude_vat**: true/false

5. **mrp_exclude_demand_charges**: true/false

6. **mrp_currency**: The currency of the MRP observation (e.g., "GHS", "KES", "NGN") — this is the LOCAL utility currency

7. **floor_ceiling_currency**: Currency of floor/ceiling rates if visible (often "USD")

8. **floor_ceiling_escalation**: Escalation mechanism if visible (e.g., "2.5% annual", "CPI", "none")

9. **mrp_clause_text**: The full OCR'd text of the formula/box

10. **tou_windows**: Time-of-use windows if any (e.g., [{"name": "peak", "start": "18:00", "end": "21:00"}, ...])

11. **discount_percentage**: Discount applied to MRP if visible (e.g., 0.21 for 21%)

12. **notes**: Any additional observations

Return as JSON only, no markdown fences."""

    response = client.messages.create(
        model='claude-sonnet-4-20250514',
        max_tokens=4096,
        messages=[{
            'role': 'user',
            'content': [
                {
                    'type': 'image',
                    'source': {
                        'type': 'base64',
                        'media_type': 'image/png',
                        'data': b64,
                    },
                },
                {'type': 'text', 'text': prompt},
            ],
        }],
    )

    text = response.content[0].text
    # Strip code fences if present
    if '```json' in text:
        text = text.split('```json')[1].split('```')[0]
    elif '```' in text:
        text = text.split('```')[1].split('```')[0]

    try:
        parsed = json.loads(text.strip())
    except json.JSONDecodeError:
        log.warning(f'  Claude returned non-JSON for {sage_id}, storing raw text')
        parsed = {'mrp_clause_text': text, 'parse_error': True}

    # Cache result
    cache_path.write_text(json.dumps(parsed, indent=2))
    return parsed


def _parse_mrp_data_rows(ws, sage_id: str) -> List[Dict[str, Any]]:
    """Parse monthly MRP data rows from a worksheet.

    Handles three layouts:
    1. Single section (standard): one ZDAT column in the header row
    2. Side-by-side dual sections (GBL01, JAB01): Grid cols left + Generator cols right, same rows
    3. Stacked sections: separate header rows for each section

    Returns list of dicts with keys: period_year, period_month, components (dict), total_per_kwh, mrp_type.
    """
    rows = []
    max_row = ws.max_row or 200
    max_col = ws.max_column or 20

    # Read all data into a 2D list for easier processing
    data = []
    for r in range(1, min(max_row + 1, 500)):
        row_vals = []
        for c in range(1, min(max_col + 1, 30)):
            cell = ws.cell(row=r, column=c)
            row_vals.append(cell.value)
        data.append(row_vals)

    if not data:
        return rows

    # Find header row containing ZDAT and identify column sections
    zdat_row_idx = None
    for i, row in enumerate(data):
        row_str = ' '.join(str(v or '').strip().upper() for v in row[:20])
        if 'ZDAT' in row_str:
            zdat_row_idx = i
            break

    if zdat_row_idx is None:
        # Try finding rows where a cell looks like a date
        for i, row in enumerate(data):
            for c_val in row[:10]:
                if c_val and _try_parse_date(c_val):
                    if i > 0:
                        zdat_row_idx = i - 1
                    break
            if zdat_row_idx is not None:
                break

    if zdat_row_idx is None:
        log.warning(f'  No header row found in sheet {sage_id}')
        return rows

    header = [str(v or '').strip().upper() for v in data[zdat_row_idx]]

    # Find all ZDAT/date columns — each one starts a section
    zdat_cols = []
    for ci, h in enumerate(header):
        if 'ZDAT' in h or h == 'PERIOD' or h == 'DATE':
            zdat_cols.append(ci)

    if not zdat_cols:
        log.warning(f'  No ZDAT column found in header row {zdat_row_idx} of {sage_id}')
        return rows

    # Also read the descriptive header row (one row below ZDAT) for component names
    desc_row_idx = zdat_row_idx + 1
    desc_header = []
    if desc_row_idx < len(data):
        desc_header = [str(v or '').strip().upper() for v in data[desc_row_idx]]

    # Build sections — each section spans from its ZDAT col to the next section's start
    sections = []
    for si, zdat_col in enumerate(zdat_cols):
        next_start = zdat_cols[si + 1] if si + 1 < len(zdat_cols) else len(header)

        # Determine mrp_type from section title (row above ZDAT)
        mrp_type = 'grid'
        # Check title row (row before ZDAT header for section labels)
        for check_row_idx in range(max(0, zdat_row_idx - 2), zdat_row_idx):
            for ci in range(zdat_col, min(next_start, len(data[check_row_idx]) if check_row_idx < len(data) else 0)):
                cell_str = str(data[check_row_idx][ci] or '').upper()
                if 'GENERATOR' in cell_str:
                    mrp_type = 'generator'
                    break

        if sage_id in GENERATOR_ONLY_PROJECTS:
            mrp_type = 'generator'

        # Build column map for this section
        col_map = {'date': zdat_col}
        for ci in range(zdat_col + 1, next_start):
            h = header[ci] if ci < len(header) else ''
            d = desc_header[ci] if ci < len(desc_header) else ''
            h_combined = (h + ' ' + d).lower()

            if 'zpritot' in h_combined or ('price total' in h_combined and 'diesel' not in h_combined):
                # For generator sections, the "Price total" after efficiency/surcharge is the computed price
                if mrp_type == 'generator' and 'total' in col_map:
                    col_map['generator_total'] = ci
                else:
                    col_map['total'] = ci
            elif 'energy' in h_combined and 'available' not in h_combined:
                col_map['energy_charge'] = ci
            elif 'levy' in h_combined:
                col_map['levy'] = ci
            elif 'lighting' in h_combined or 'light' in h_combined:
                col_map['lighting'] = ci
            elif 'subsidy' in h_combined:
                col_map['subsidy'] = ci
            elif 'fuel' in h_combined or 'diesel' in h_combined:
                col_map['fuel_price'] = ci
            elif 'service' in h_combined:
                col_map['service_charge'] = ci
            elif 'grid' in h_combined and 'reference' in h_combined:
                col_map['grid_reference'] = ci
            elif 'efficiency' in h_combined:
                col_map['generator_efficiency'] = ci
            elif 'surcharge' in h_combined:
                col_map['generator_surcharge'] = ci

        sections.append({
            'zdat_col': zdat_col,
            'col_map': col_map,
            'mrp_type': mrp_type,
        })

    log.info(f'  Found {len(sections)} section(s) in {sage_id}: '
             f'{[(s["mrp_type"], s["zdat_col"]) for s in sections]}')

    # Parse data rows — start 2 rows after ZDAT header (skip descriptive row)
    data_start = zdat_row_idx + 2
    for di in range(data_start, len(data)):
        drow = data[di]

        for section in sections:
            cm = section['col_map']
            date_col = cm['date']
            date_val = drow[date_col] if date_col < len(drow) else None
            parsed_date = _try_parse_date(date_val)
            if not parsed_date:
                continue

            components = {}
            for comp_name, ci in cm.items():
                if comp_name == 'date':
                    continue
                val = drow[ci] if ci < len(drow) else None
                if val is not None:
                    try:
                        components[comp_name] = float(val)
                    except (ValueError, TypeError):
                        pass

            # Get total — for generator sections, prefer generator_total if available
            total = None
            if section['mrp_type'] == 'generator' and 'generator_total' in components:
                total = components.pop('generator_total')
                components.pop('total', None)
            else:
                total = components.pop('total', None)

            if total is None:
                # Sum non-metadata components
                skip_keys = {'generator_efficiency', 'generator_surcharge'}
                total = sum(v for k, v in components.items()
                           if v is not None and k not in skip_keys)

            if total == 0 and not components:
                continue

            rows.append({
                'period_year': parsed_date.year,
                'period_month': parsed_date.month,
                'components': components,
                'total_per_kwh': total,
                'mrp_type': section['mrp_type'],
            })

    return rows


def _try_parse_date(val) -> Optional[date]:
    """Try to parse a cell value as a date."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%Y%m'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Try Excel serial date
    try:
        serial = float(s)
        if 30000 < serial < 60000:
            return (datetime(1899, 12, 30) + timedelta(days=serial)).date()
    except (ValueError, TypeError):
        pass
    return None


def phase_b(dry_run: bool, use_cache: bool = True, project_filter: Optional[str] = None) -> Dict[str, Any]:
    """OCR MRP formula screenshots and parse monthly MRP data."""
    log.info('=== Phase B: MRP Formula OCR + Monthly Data ===')
    result = {
        'phase': 'B',
        'sheets_processed': 0,
        'images_extracted': 0,
        'ocr_results': {},
        'tariffs_updated': 0,
        'reference_prices_inserted': 0,
        'reference_prices_skipped': 0,
        'consistency_checks': [],
        'errors': [],
    }

    if not MRP_XLSX.exists():
        result['errors'].append(f'MRP XLSX not found: {MRP_XLSX}')
        log.error(result['errors'][-1])
        return result

    # ---- Load DB state ----
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '300s'")

        # Projects
        cur.execute("""
            SELECT p.id, p.sage_id, p.country, p.cod_date::text as cod_date,
                   p.technical_specs
            FROM project p WHERE p.organization_id = %s
        """, (ORG_ID,))
        projects = {r['sage_id']: dict(r) for r in cur.fetchall()}

        # Clause tariffs (MRP-relevant, is_current=true)
        # Note: some clause_tariff rows have organization_id = NULL
        cur.execute("""
            SELECT ct.id, ct.project_id, ct.currency_id, ct.market_ref_currency_id,
                   ct.logic_parameters, ct.name,
                   p.sage_id
            FROM clause_tariff ct
            JOIN project p ON p.id = ct.project_id
            WHERE ct.is_current = true
              AND p.organization_id = %s
        """, (ORG_ID,))
        tariffs_by_sage: Dict[str, List[Dict]] = {}
        for r in cur.fetchall():
            sid = r['sage_id']
            tariffs_by_sage.setdefault(sid, []).append(dict(r))

        # Currencies
        cur.execute("SELECT id, code FROM currency")
        currencies = {r['code']: r['id'] for r in cur.fetchall()}
        currency_by_id = {r['id']: r['code'] for r in cur.fetchall()}
        # Re-query for id→code since fetchall consumed the cursor
        cur.execute("SELECT id, code FROM currency")
        for r in cur.fetchall():
            currency_by_id[r['id']] = r['code']

        # Existing reference_price for dedup/comparison
        cur.execute("""
            SELECT project_id, observation_type, period_start,
                   calculated_mrp_per_kwh, currency_id
            FROM reference_price
            WHERE organization_id = %s AND observation_type = 'monthly'
        """, (ORG_ID,))
        existing_rp: Dict[Tuple, Dict] = {}
        for r in cur.fetchall():
            key = (r['project_id'], str(r['period_start']))
            existing_rp[key] = dict(r)

        # Billing periods for operating_year calculation
        cur.execute("SELECT id, start_date, end_date FROM billing_period ORDER BY start_date")
        bp_by_ym = {}
        for bp in cur.fetchall():
            if bp['start_date']:
                bp_by_ym[(bp['start_date'].year, bp['start_date'].month)] = dict(bp)

    # ---- Open XLSX ----
    log.info(f'Opening MRP workbook: {MRP_XLSX}')
    wb = openpyxl.load_workbook(str(MRP_XLSX), data_only=True)
    sheet_names = wb.sheetnames
    log.info(f'Sheets: {sheet_names}')

    tariff_updates = []
    rp_inserts = []

    for sheet_name in sheet_names:
        # Resolve sage_id from sheet name
        sage_id = None
        for alias, canonical in SAGE_ALIASES.items():
            if alias.upper() in sheet_name.upper():
                sage_id = canonical
                break
        if not sage_id:
            log.info(f'  Skipping sheet "{sheet_name}" (no sage_id match)')
            continue

        if project_filter and sage_id != project_filter:
            continue

        log.info(f'Processing sheet "{sheet_name}" → {sage_id}')
        result['sheets_processed'] += 1
        proj = projects.get(sage_id)
        if not proj:
            result['errors'].append(f'Project {sage_id} not found in DB')
            continue

        # ---- B1: Extract & OCR formula images ----
        ws = wb[sheet_name]
        images = _extract_images_from_sheet(wb, sheet_name)
        result['images_extracted'] += len(images)

        ocr_data = None
        for idx, img_bytes in enumerate(images):
            # Save image
            img_path = IMAGE_DIR / f'{sage_id}_{idx}.png'
            img_path.write_bytes(img_bytes)
            log.info(f'  Saved image: {img_path.name} ({len(img_bytes)} bytes)')

            # OCR
            ocr = _ocr_image_claude(img_bytes, sage_id, use_cache=use_cache)
            result['ocr_results'][f'{sage_id}_{idx}'] = ocr

            if not ocr.get('parse_error'):
                ocr_data = ocr  # Use last successfully parsed OCR

        # ---- B2: Update clause_tariff MRP rules ----
        if ocr_data and sage_id not in ALREADY_POPULATED:
            tariffs = tariffs_by_sage.get(sage_id, [])
            if not tariffs:
                result['errors'].append(f'{sage_id}: no clause_tariff rows found')
            else:
                # Determine MRP currency from OCR or country
                mrp_currency_code = ocr_data.get('mrp_currency')
                if not mrp_currency_code:
                    mrp_currency_code = COUNTRY_CURRENCY.get(proj.get('country'), '')
                mrp_currency_id = currencies.get(mrp_currency_code)

                for tariff in tariffs:
                    lp = tariff.get('logic_parameters') or {}
                    if lp.get('mrp_method'):
                        continue  # Already populated

                    new_lp = {
                        **lp,
                        'mrp_method': ocr_data.get('mrp_method', 'utility_variable_charges_tou'),
                        'mrp_clause_text': ocr_data.get('mrp_clause_text', ''),
                        'mrp_included_components': ocr_data.get('mrp_included_components', []),
                        'mrp_excluded_components': ocr_data.get('mrp_excluded_components', []),
                        'mrp_exclude_vat': ocr_data.get('mrp_exclude_vat', True),
                        'mrp_exclude_demand_charges': ocr_data.get('mrp_exclude_demand_charges', True),
                    }
                    if ocr_data.get('floor_ceiling_currency'):
                        new_lp['floor_ceiling_currency'] = ocr_data['floor_ceiling_currency']
                    if ocr_data.get('floor_ceiling_escalation'):
                        new_lp['floor_ceiling_escalation'] = ocr_data['floor_ceiling_escalation']
                    if ocr_data.get('tou_windows'):
                        new_lp['tou_windows'] = ocr_data['tou_windows']
                    if ocr_data.get('discount_percentage'):
                        new_lp['discount_percentage'] = ocr_data['discount_percentage']

                    tariff_updates.append({
                        'tariff_id': tariff['id'],
                        'sage_id': sage_id,
                        'logic_parameters': new_lp,
                        'market_ref_currency_id': mrp_currency_id,
                    })
                    log.info(f'  Will update clause_tariff {tariff["id"]} for {sage_id}: '
                             f'mrp_method={new_lp["mrp_method"]}, currency={mrp_currency_code}')

        elif ocr_data and sage_id in ALREADY_POPULATED:
            # Cross-validate existing tariff data
            tariffs = tariffs_by_sage.get(sage_id, [])
            for tariff in tariffs:
                lp = tariff.get('logic_parameters') or {}
                existing_method = lp.get('mrp_method')
                ocr_method = ocr_data.get('mrp_method')
                if existing_method and ocr_method and existing_method != ocr_method:
                    result['consistency_checks'].append({
                        'sage_id': sage_id,
                        'level': 'warning',
                        'message': f'MRP method mismatch: DB={existing_method}, OCR={ocr_method}',
                    })

        # ---- B3: Parse monthly MRP data rows ----
        mrp_rows = _parse_mrp_data_rows(ws, sage_id)
        log.info(f'  Parsed {len(mrp_rows)} MRP data rows from {sage_id}')

        project_id = proj['id']
        # Determine MRP currency_id
        tariffs = tariffs_by_sage.get(sage_id, [])
        mrp_currency_id = None
        for t in tariffs:
            if t.get('market_ref_currency_id'):
                mrp_currency_id = t['market_ref_currency_id']
                break
        if not mrp_currency_id:
            # Fall back to country currency
            cc = COUNTRY_CURRENCY.get(proj.get('country'))
            if cc:
                mrp_currency_id = currencies.get(cc)

        # Calculate operating year from COD
        cod_str = proj.get('cod_date')
        cod_date_val = None
        if cod_str:
            try:
                cod_date_val = datetime.strptime(cod_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        for mrp_row in mrp_rows:
            y, m = mrp_row['period_year'], mrp_row['period_month']
            period_start = date(y, m, 1)
            _, last_day = monthrange(y, m)
            period_end = date(y, m, last_day)

            # Operating year
            oy = 1
            if cod_date_val:
                delta_years = (period_start.year - cod_date_val.year)
                if period_start.month < cod_date_val.month:
                    delta_years -= 1
                oy = max(1, delta_years + 1)

            # Check for existing reference_price
            key = (project_id, str(period_start))
            if key in existing_rp:
                existing = existing_rp[key]
                if sage_id in ALREADY_POPULATED:
                    # Compare
                    ex_val = float(existing.get('calculated_mrp_per_kwh') or 0)
                    new_val = mrp_row['total_per_kwh']
                    if ex_val > 0 and new_val > 0:
                        diff_pct = abs(ex_val - new_val) / ex_val * 100
                        if diff_pct > 5:
                            result['consistency_checks'].append({
                                'sage_id': sage_id,
                                'period': str(period_start),
                                'level': 'warning',
                                'message': f'MRP value diff {diff_pct:.1f}%: DB={ex_val:.6f}, XLSX={new_val:.6f}',
                            })
                result['reference_prices_skipped'] += 1
                continue

            rp_inserts.append({
                'project_id': project_id,
                'organization_id': ORG_ID,
                'operating_year': oy,
                'period_start': period_start,
                'period_end': period_end,
                'calculated_mrp_per_kwh': mrp_row['total_per_kwh'],
                'currency_id': mrp_currency_id,
                'verification_status': 'estimated',
                'observation_type': 'monthly',
                'source_metadata': json.dumps({
                    'source_file': 'Sage Contract Extracts market Ref pricing data.xlsx',
                    'sheet_name': sheet_name,
                    'mrp_type': mrp_row['mrp_type'],
                    'tariff_components': mrp_row['components'],
                    'extraction_date': date.today().isoformat(),
                }, default=str),
            })

    wb.close()

    # Deduplicate reference_price inserts by (project_id, period_start).
    # For dual-section projects (grid + generator), merge into one row:
    #   - Use grid total as calculated_mrp_per_kwh (primary MRP)
    #   - Store both sections in source_metadata
    deduped_rp: Dict[Tuple, Dict] = {}
    for rp in rp_inserts:
        key = (rp['project_id'], str(rp['period_start']))
        sm = json.loads(rp['source_metadata'])
        mrp_type = sm.get('mrp_type', 'grid')

        if key not in deduped_rp:
            deduped_rp[key] = rp
        else:
            existing = deduped_rp[key]
            existing_sm = json.loads(existing['source_metadata'])

            # Prefer grid over generator for primary MRP value
            if mrp_type == 'grid' and existing_sm.get('mrp_type') != 'grid':
                # Swap: new grid becomes primary, old generator goes to metadata
                sm['generator_total_per_kwh'] = existing['calculated_mrp_per_kwh']
                sm['generator_components'] = existing_sm.get('tariff_components', {})
                rp['source_metadata'] = json.dumps(sm, default=str)
                deduped_rp[key] = rp
            else:
                # Keep existing as primary, store new in metadata
                existing_sm[f'{mrp_type}_total_per_kwh'] = rp['calculated_mrp_per_kwh']
                existing_sm[f'{mrp_type}_components'] = sm.get('tariff_components', {})
                existing['source_metadata'] = json.dumps(existing_sm, default=str)

    rp_inserts = list(deduped_rp.values())

    log.info(f'Phase B totals: {len(tariff_updates)} tariff updates, {len(rp_inserts)} reference_price inserts (after dedup)')

    if dry_run:
        result['tariffs_updated'] = len(tariff_updates)
        result['reference_prices_inserted'] = len(rp_inserts)
        log.info(f'[DRY RUN] Would update {len(tariff_updates)} tariffs, insert {len(rp_inserts)} reference_prices')
        return result

    # ---- Write to DB ----
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '300s'")

        # B2: Update clause_tariff
        for upd in tariff_updates:
            cur.execute("""
                UPDATE clause_tariff
                SET logic_parameters = %s,
                    market_ref_currency_id = %s
                WHERE id = %s
            """, (json.dumps(upd['logic_parameters'], default=str),
                  upd['market_ref_currency_id'],
                  upd['tariff_id']))
        result['tariffs_updated'] = len(tariff_updates)
        log.info(f'Updated {len(tariff_updates)} clause_tariff rows')

        # B3: Insert reference_price
        if rp_inserts:
            cols = [
                'project_id', 'organization_id', 'operating_year',
                'period_start', 'period_end', 'calculated_mrp_per_kwh',
                'currency_id', 'verification_status', 'observation_type',
                'source_metadata',
            ]
            insert_sql = f"""
                INSERT INTO reference_price ({', '.join(cols)})
                VALUES %s
                ON CONFLICT (project_id, observation_type, period_start) DO UPDATE
                SET calculated_mrp_per_kwh = EXCLUDED.calculated_mrp_per_kwh,
                    currency_id = EXCLUDED.currency_id,
                    source_metadata = EXCLUDED.source_metadata,
                    updated_at = NOW()
            """
            values = [tuple(r[c] for c in cols) for r in rp_inserts]
            execute_values(cur, insert_sql, values, page_size=100)
            result['reference_prices_inserted'] = cur.rowcount
            log.info(f'Inserted {cur.rowcount} reference_price rows')

    return result


# =========================================================================
# Phase C: Plant Performance Enrichment
# =========================================================================

def phase_c(dry_run: bool, project_filter: Optional[str] = None) -> Dict[str, Any]:
    """Enrich plant_performance from meter_aggregate + production_forecast."""
    log.info('=== Phase C: Plant Performance Enrichment ===')
    result = {
        'phase': 'C',
        'rows_updated': 0,
        'rows_inserted': 0,
        'projects_covered': set(),
        'info_gaps': [],
        'errors': [],
    }

    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '300s'")

        # Aggregate meter data per project per month
        cur.execute("""
            SELECT
                p.sage_id,
                c.project_id,
                ma.billing_period_id,
                bp.start_date AS billing_month,
                SUM(CASE WHEN cl.energy_category = 'metered' THEN ma.total_production ELSE 0 END) AS total_metered_kwh,
                SUM(CASE WHEN cl.energy_category = 'available' THEN COALESCE(ma.available_energy_kwh, ma.total_production) ELSE 0 END) AS total_available_kwh
            FROM meter_aggregate ma
            JOIN contract_line cl ON cl.id = ma.contract_line_id
            JOIN contract c ON c.id = cl.contract_id
            JOIN project p ON p.id = c.project_id
            JOIN billing_period bp ON bp.id = ma.billing_period_id
            WHERE ma.organization_id = %s AND ma.period_type = 'monthly'
            GROUP BY p.sage_id, c.project_id, ma.billing_period_id, bp.start_date
            ORDER BY c.project_id, bp.start_date
        """, (ORG_ID,))
        meter_agg = cur.fetchall()

        if not meter_agg:
            log.info('No meter_aggregate data found — skipping Phase C')
            result['info_gaps'].append('No meter_aggregate data available')
            return result

        # Load production_forecast
        cur.execute("""
            SELECT id, project_id, forecast_month, forecast_energy_kwh,
                   forecast_ghi_irradiance, forecast_pr
            FROM production_forecast
            WHERE organization_id = %s
        """, (ORG_ID,))
        forecasts: Dict[Tuple, Dict] = {}
        for r in cur.fetchall():
            key = (r['project_id'], str(r['forecast_month']))
            forecasts[key] = dict(r)

        # Load existing plant_performance
        cur.execute("""
            SELECT project_id, billing_month
            FROM plant_performance
            WHERE organization_id = %s
        """, (ORG_ID,))
        existing_pp = {(r['project_id'], str(r['billing_month'])) for r in cur.fetchall()}

        # COD dates for operating_year
        cur.execute("""
            SELECT id, sage_id, cod_date::text as cod_date
            FROM project WHERE organization_id = %s
        """, (ORG_ID,))
        proj_cod = {r['id']: r for r in cur.fetchall()}

        upserts = []
        for row in meter_agg:
            sage_id = row['sage_id']
            project_id = row['project_id']
            billing_month = row['billing_month']

            if project_filter and sage_id != project_filter:
                continue

            result['projects_covered'].add(sage_id)

            bm_str = str(billing_month)
            forecast_key = (project_id, bm_str)
            forecast = forecasts.get(forecast_key)

            energy_comparison = None
            if forecast and forecast.get('forecast_energy_kwh') and float(forecast['forecast_energy_kwh']) > 0:
                total_metered = float(row['total_metered_kwh'] or 0)
                if total_metered > 0:
                    energy_comparison = total_metered / float(forecast['forecast_energy_kwh'])

            # Operating year
            oy = None
            pinfo = proj_cod.get(project_id)
            if pinfo and pinfo.get('cod_date'):
                try:
                    cod = datetime.strptime(pinfo['cod_date'], '%Y-%m-%d').date()
                    delta = (billing_month.year - cod.year)
                    if billing_month.month < cod.month:
                        delta -= 1
                    oy = max(1, delta + 1)
                except (ValueError, TypeError):
                    pass

            bp_id = row['billing_period_id']
            pf_id = forecast['id'] if forecast else None

            upserts.append({
                'project_id': project_id,
                'organization_id': ORG_ID,
                'production_forecast_id': pf_id,
                'billing_period_id': bp_id,
                'billing_month': billing_month,
                'operating_year': oy,
                'energy_comparison': energy_comparison,
            })

        log.info(f'Phase C: {len(upserts)} plant_performance rows to upsert across {len(result["projects_covered"])} projects')

        if not forecast:
            result['info_gaps'].append('irr_comparison and pr_comparison require irradiance data — not available in CSV')

        if dry_run:
            result['rows_inserted'] = len(upserts)
            result['projects_covered'] = sorted(result['projects_covered'])
            log.info(f'[DRY RUN] Would upsert {len(upserts)} plant_performance rows')
            return result

        # Upsert
        for u in upserts:
            pp_key = (u['project_id'], str(u['billing_month']))
            if pp_key in existing_pp:
                # Update only energy_comparison if we have it
                if u['energy_comparison'] is not None:
                    cur.execute("""
                        UPDATE plant_performance
                        SET energy_comparison = %s,
                            production_forecast_id = COALESCE(%s, production_forecast_id),
                            updated_at = NOW()
                        WHERE project_id = %s AND billing_month = %s
                    """, (u['energy_comparison'], u['production_forecast_id'],
                          u['project_id'], u['billing_month']))
                    result['rows_updated'] += 1
            else:
                cur.execute("""
                    INSERT INTO plant_performance
                        (project_id, organization_id, production_forecast_id,
                         billing_period_id, billing_month, operating_year,
                         energy_comparison)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (project_id, billing_month) DO UPDATE
                    SET energy_comparison = EXCLUDED.energy_comparison,
                        updated_at = NOW()
                """, (u['project_id'], u['organization_id'],
                      u['production_forecast_id'], u['billing_period_id'],
                      u['billing_month'], u['operating_year'],
                      u['energy_comparison']))
                result['rows_inserted'] += 1

        log.info(f'Phase C: {result["rows_inserted"]} inserted, {result["rows_updated"]} updated')

    result['projects_covered'] = sorted(result['projects_covered'])
    return result


# =========================================================================
# Main
# =========================================================================

def main():
    parser = argparse.ArgumentParser(description='Step 9 & 10: MRP + Meter Population')
    parser.add_argument('--dry-run', action='store_true', help='Do not write to DB')
    parser.add_argument('--no-cache', action='store_true', help='Force re-OCR (ignore cache)')
    parser.add_argument('--phase', type=str, default='all',
                        choices=['A', 'B', 'C', 'all'],
                        help='Run specific phase (default: all)')
    parser.add_argument('--project', type=str, default=None,
                        help='Process single project by SAGE ID')
    args = parser.parse_args()

    mode = 'DRY RUN' if args.dry_run else 'LIVE'
    log.info(f'Step 9 & 10: MRP + Meter Population — {mode}')
    if args.project:
        log.info(f'  Filtering to project: {args.project}')

    init_connection_pool()

    phases_result = {}
    try:
        if args.phase in ('A', 'all'):
            phases_result['A'] = phase_a(args.dry_run, args.project)

        if args.phase in ('B', 'all'):
            phases_result['B'] = phase_b(args.dry_run, use_cache=not args.no_cache,
                                          project_filter=args.project)

        if args.phase in ('C', 'all'):
            phases_result['C'] = phase_c(args.dry_run, args.project)

    finally:
        close_connection_pool()

    # ---- Gate checks ----
    gates = []
    pa = phases_result.get('A', {})
    pb = phases_result.get('B', {})

    if pa:
        inserted = pa.get('rows_inserted', 0)
        parsed = pa.get('rows_parsed', 0)
        unresolved_cl = len(pa.get('unresolved_contract_lines', []))
        fk_rate = ((parsed - unresolved_cl) / parsed * 100) if parsed > 0 else 0

        gates.append({
            'gate': 'Meter readings inserted',
            'expected': '> 550 of 604',
            'actual': inserted,
            'status': 'pass' if inserted >= 550 else 'warning',
        })
        gates.append({
            'gate': 'FK resolution rate',
            'expected': '> 95%',
            'actual': f'{fk_rate:.1f}%',
            'status': 'pass' if fk_rate >= 95 else 'warning',
        })

    if pb:
        sheets = pb.get('sheets_processed', 0)
        tariffs = pb.get('tariffs_updated', 0)
        rp = pb.get('reference_prices_inserted', 0)
        consistency = pb.get('consistency_checks', [])
        critical = [c for c in consistency if c.get('level') == 'critical']

        gates.append({
            'gate': 'MRP formula OCR\'d',
            'expected': '>= 7 project tabs',
            'actual': sheets,
            'status': 'pass' if sheets >= 7 else 'warning',
        })
        gates.append({
            'gate': 'MRP rules populated in clause_tariff',
            'expected': '>= 7 tariffs',
            'actual': tariffs,
            'status': 'pass' if tariffs >= 7 else 'warning',
        })
        gates.append({
            'gate': 'MRP observations inserted',
            'expected': '>= 300',
            'actual': rp,
            'status': 'pass' if rp >= 300 else 'warning',
        })
        gates.append({
            'gate': 'No critical discrepancies',
            'expected': '0 critical',
            'actual': len(critical),
            'status': 'pass' if len(critical) == 0 else 'critical',
        })

    # Overall status
    statuses = [g['status'] for g in gates]
    if 'critical' in statuses:
        status = 'critical'
    elif 'warning' in statuses:
        status = 'warnings'
    else:
        status = 'pass'

    # ---- Build report ----
    report = {
        'step': 9,
        'step_name': 'MRP + Meter Population & Plant Performance',
        'mode': mode,
        'status': status,
        'run_date': date.today().isoformat(),
        'phases': phases_result,
        'gate_checks': gates,
    }

    report_path = REPORT_DIR / f'step9_{date.today().isoformat()}.json'
    with open(report_path, 'w') as f:
        json.dump(report, f, indent=2, default=str)
    log.info(f'Report saved: {report_path}')
    log.info(f'Overall status: {status}')

    # Print summary
    print(f'\n{"="*60}')
    print(f'Step 9 & 10 — {mode}')
    print(f'{"="*60}')
    for g in gates:
        icon = '✓' if g['status'] == 'pass' else '⚠' if g['status'] == 'warning' else '✗'
        print(f'  {icon} {g["gate"]}: {g["actual"]} (expected {g["expected"]})')
    print(f'\nOverall: {status}')
    print(f'Report: {report_path}')


if __name__ == '__main__':
    main()
