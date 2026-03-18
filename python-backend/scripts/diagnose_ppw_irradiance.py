#!/usr/bin/env python3
"""
Diagnostic: inspect PPW project tabs for GHI/POA irradiance and PR columns.
Reports what each project tab has in:
  1. Monthly Allocation section (rows 8-20)
  2. Technical Model header columns
  3. Sample values for Jan OY1

Usage:
    cd python-backend
    python scripts/diagnose_ppw_irradiance.py
"""

import os
import sys
from pathlib import Path
from datetime import date, datetime
from typing import Any, Optional

script_dir = Path(__file__).resolve().parent
project_root = script_dir.parent
sys.path.insert(0, str(project_root))

from services.onboarding.parsers.plant_performance_parser import (
    TAB_NAME_TO_SAGE_ID,
    TECH_MODEL_COLUMNS,
    _safe_float,
    _parse_date,
)

WORKBOOK_PATH = os.path.join(
    project_root.parent, "CBE_data_extracts", "Operations Plant Performance Workbook.xlsx"
)


def main():
    import openpyxl

    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)

    print("=" * 120)
    print(f"{'SAGE_ID':<10} {'Tab':<20} {'Tech GHI col?':<15} {'Tech POA col?':<15} {'Tech PR GHI?':<14} {'Tech PR POA?':<14} {'Alloc GHI?':<12} {'Alloc POA?':<12} {'Sample GHI':<12} {'Sample POA':<12}")
    print("=" * 120)

    for sheet_name in wb.sheetnames:
        sage_id = TAB_NAME_TO_SAGE_ID.get(sheet_name)
        if not sage_id:
            continue
        # Skip summary tabs
        if "summary" in sheet_name.lower() or "waterfall" in sheet_name.lower():
            continue

        rows = [list(r) for r in wb[sheet_name].iter_rows(values_only=True)]
        if len(rows) < 10:
            continue

        # ── Check Technical Model header ──
        tech_ghi_col = False
        tech_poa_col = False
        tech_pr_ghi = False
        tech_pr_poa = False
        sample_ghi = None
        sample_poa = None
        tech_header_idx = None

        for i, row in enumerate(rows):
            if not row:
                continue
            cells_lower = [str(c).strip().lower() if c is not None else "" for c in row]
            has_date = any(c in ("date", "month") for c in cells_lower)
            has_oy = "oy" in cells_lower
            if not has_date or not has_oy:
                continue

            col_map = {}
            for j, cell_lower in enumerate(cells_lower):
                for pattern, field_name in TECH_MODEL_COLUMNS:
                    if pattern in cell_lower:
                        if field_name not in col_map.values():
                            col_map[j] = field_name
                        break

            if len(col_map) >= 3:
                tech_header_idx = i
                for j, field in col_map.items():
                    if field == "forecast_ghi_wm2":
                        tech_ghi_col = True
                    elif field == "forecast_poa_wm2":
                        tech_poa_col = True
                    elif field == "forecast_pr":
                        tech_pr_ghi = True
                    elif field == "forecast_pr_poa":
                        tech_pr_poa = True

                # Print the actual header cells for irradiance columns
                irr_headers = []
                for j, field in col_map.items():
                    if "ghi" in field or "poa" in field or "pr" in field.lower():
                        irr_headers.append(f"  col {j}: '{row[j]}' -> {field}")

                # Get first data row sample
                if i + 1 < len(rows):
                    data_row = rows[i + 1]
                    for j, field in col_map.items():
                        if field == "forecast_ghi_wm2" and j < len(data_row):
                            sample_ghi = _safe_float(data_row[j])
                        elif field == "forecast_poa_wm2" and j < len(data_row):
                            sample_poa = _safe_float(data_row[j])
                break

        # ── Check Monthly Allocation section ──
        alloc_ghi = False
        alloc_poa = False
        for i, row in enumerate(rows[:30]):
            if not row:
                continue
            for j, cell in enumerate(row):
                if cell and isinstance(cell, str):
                    cl = cell.lower().strip()
                    if "ghi" in cl:
                        alloc_ghi = True
                    if "poa" in cl:
                        alloc_poa = True

        yn = lambda b: "YES" if b else "no"
        print(
            f"{sage_id:<10} {sheet_name:<20} "
            f"{yn(tech_ghi_col):<15} {yn(tech_poa_col):<15} "
            f"{yn(tech_pr_ghi):<14} {yn(tech_pr_poa):<14} "
            f"{yn(alloc_ghi):<12} {yn(alloc_poa):<12} "
            f"{str(sample_ghi):<12} {str(sample_poa):<12}"
        )

        # Print detailed irradiance column headers if they exist
        if tech_header_idx is not None:
            header_row = rows[tech_header_idx]
            for j, cell in enumerate(header_row):
                if cell is None:
                    continue
                cl = str(cell).lower()
                if any(kw in cl for kw in ("ghi", "poa", "pr ")):
                    print(f"    col {j}: '{cell}'")

    wb.close()
    print("\n" + "=" * 120)


if __name__ == "__main__":
    main()
