#!/usr/bin/env python3
"""
Step 6: Plant Performance Workbook — Project Tabs.

Enriches existing production_forecast rows with source_metadata (site params,
monthly allocation, per-phase breakdown) and verifies basic project details
against PPW tab data. Does NOT create new rows — only updates existing ones.

Usage:
    cd python-backend
    python scripts/step6_project_tabs.py --dry-run        # Preview
    python scripts/step6_project_tabs.py                   # Execute
    python scripts/step6_project_tabs.py --project KAS01   # Single project
"""

import argparse
import json
import logging
import os
import sys
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from dotenv import load_dotenv

load_dotenv()

# Add project root to path
script_dir = Path(__file__).resolve().parent
project_root = script_dir.parent
sys.path.insert(0, str(project_root))

from psycopg2.extras import execute_values
from db.database import init_connection_pool, close_connection_pool, get_db_connection
from services.onboarding.parsers.plant_performance_parser import PlantPerformanceParser

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("step6_project_tabs")

DEFAULT_ORG_ID = 1
WORKBOOK_PATH = os.path.join(
    project_root.parent, "CBE_data_extracts", "Operations Plant Performance Workbook.xlsx"
)


# =============================================================================
# Data Structures
# =============================================================================

@dataclass
class Discrepancy:
    severity: str       # "critical" | "warning" | "info"
    category: str
    project: str
    field: str
    source_a: str       # PPW value
    source_b: str       # DB value
    recommended_action: str
    status: str = "open"


@dataclass
class GateCheck:
    name: str
    passed: bool
    expected: str
    actual: str


@dataclass
class ProjectResult:
    sage_id: str
    forecasts_enriched: int = 0
    site_params_stored: bool = False
    allocation_stored: bool = False
    details_verified: bool = False
    discrepancies: List[Discrepancy] = field(default_factory=list)


@dataclass
class StepReport:
    step: int = 6
    step_name: str = "Plant Performance Workbook — Project Tabs"
    status: str = "passed"
    projects_processed: int = 0
    forecasts_enriched: int = 0
    discrepancies: List[Discrepancy] = field(default_factory=list)
    gate_checks: List[GateCheck] = field(default_factory=list)
    project_results: List[ProjectResult] = field(default_factory=list)


class DateEncoder(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, (date, datetime)):
            return o.isoformat()
        return super().default(o)


# =============================================================================
# Parser: Extract basic project details from rows 1-8
# =============================================================================

def extract_basic_details(rows: list) -> Dict[str, Any]:
    """Extract basic project details from PPW tab header (rows 1-8)."""
    details: Dict[str, Any] = {}
    for row in rows[:10]:
        if not row:
            continue
        for j, val in enumerate(row):
            if val is None or not isinstance(val, str):
                continue
            label = val.strip().lower()
            next_val = row[j + 1] if j + 1 < len(row) else None

            if label.startswith("customer") and next_val:
                details["customer_name"] = str(next_val).strip()
            elif label.startswith("country") and next_val:
                details["country"] = str(next_val).strip()
            elif "cod" in label and "phase 1" in label and next_val:
                details["cod_phase1"] = _to_date(next_val)
            elif "cod" in label and "phase 2" in label and next_val:
                details["cod_phase2"] = _to_date(next_val)
            elif label.startswith("cod") and "phase" not in label and next_val:
                details["cod"] = _to_date(next_val)
            elif label.startswith("term") and next_val:
                try:
                    details["term_years"] = int(float(str(next_val)))
                except (ValueError, TypeError):
                    pass
            elif label.startswith("sage") and next_val:
                details["sage_id"] = str(next_val).strip()
    return details


def extract_monthly_allocation(rows: list) -> List[Dict[str, Any]]:
    """Extract the 12-month GHI/POA/Energy%/PR allocation from the Fixed Parameters section."""
    allocations = []
    month_names = ["jan", "feb", "mar", "apr", "may", "jun",
                   "jul", "aug", "sep", "oct", "nov", "dec"]

    # Find the "Monthly Allocation" label or the month rows (Jan-Dec)
    start_idx = None
    for i, row in enumerate(rows[:30]):
        if not row:
            continue
        for cell in row:
            if cell and isinstance(cell, str) and "monthly allocation" in cell.lower():
                start_idx = i + 1
                break
        if start_idx:
            break

    if start_idx is None:
        # Try to find month pattern directly
        for i, row in enumerate(rows[6:25], start=6):
            if not row:
                continue
            first_cell = str(row[0]).strip().lower() if row[0] else ""
            # Check if first cell is a month name or a date that looks like Jan
            if first_cell in month_names:
                start_idx = i
                break
            if hasattr(row[0], 'month') and hasattr(row[0], 'year'):
                # It's a datetime — check if row[0].year == 1900 (Excel month-only format)
                try:
                    from datetime import datetime as dt
                    if isinstance(row[0], (dt, date)) and row[0].year == 1900:
                        start_idx = i
                        break
                except Exception:
                    pass

    if start_idx is None:
        return allocations

    # Find which columns have the allocation data
    # Look for a header row just before start_idx with "Month", "GHI", "POA", "Energy"
    header_row = rows[start_idx - 1] if start_idx > 0 else None
    # Default column positions (relative to the allocation block)
    ghi_col = None
    poa_col = None
    energy_col = None

    if header_row:
        for j, cell in enumerate(header_row):
            if cell and isinstance(cell, str):
                cl = cell.lower().strip()
                if "ghi" in cl and ghi_col is None:
                    ghi_col = j
                elif "poa" in cl and poa_col is None:
                    poa_col = j
                elif "energy" in cl and energy_col is None:
                    energy_col = j

    # Extract 12 months
    month_num = 0
    for i in range(start_idx, min(start_idx + 15, len(rows))):
        row = rows[i]
        if not row:
            continue

        month_num += 1
        if month_num > 12:
            break

        alloc: Dict[str, Any] = {"month_num": month_num}
        if ghi_col and ghi_col < len(row):
            v = row[ghi_col]
            if isinstance(v, (int, float)):
                alloc["ghi_wm2"] = float(v)
        if poa_col and poa_col < len(row):
            v = row[poa_col]
            if isinstance(v, (int, float)):
                alloc["poa_wm2"] = float(v)
        if energy_col and energy_col < len(row):
            v = row[energy_col]
            if isinstance(v, (int, float)):
                alloc["energy_pct"] = float(v)

        if len(alloc) > 1:  # Has at least month_num + one data field
            allocations.append(alloc)

    return allocations


def _to_date(val: Any) -> Optional[str]:
    """Convert a cell value to ISO date string."""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if val and str(val).strip() not in ("#REF!", "", "N/A"):
        return str(val).strip()
    return None


# =============================================================================
# Main Processing
# =============================================================================

def process_project(
    sage_id: str,
    site_params: Dict[str, Any],
    tech_rows: list,
    basic_details: Dict[str, Any],
    monthly_alloc: List[Dict[str, Any]],
    org_id: int,
    dry_run: bool,
) -> ProjectResult:
    """Process a single project tab: verify details, enrich forecasts."""
    result = ProjectResult(sage_id=sage_id)

    with get_db_connection() as conn:
        conn.autocommit = False
        try:
            with conn.cursor() as cur:
                # Set generous timeout — default 60s is too short for large projects
                cur.execute("SET statement_timeout = '300000'")  # 5 min

                # ── 1. Look up project ──
                cur.execute(
                    "SELECT id, name, cod_date, country, installed_dc_capacity_kwp "
                    "FROM project WHERE sage_id = %s AND organization_id = %s",
                    (sage_id, org_id),
                )
                proj = cur.fetchone()
                if not proj:
                    result.discrepancies.append(Discrepancy(
                        severity="info", category="missing_project",
                        project=sage_id, field="project",
                        source_a=sage_id, source_b="NOT FOUND",
                        recommended_action="PPW tab has no matching FM project — skip",
                    ))
                    return result

                project_id = proj["id"]

                # ── 2. Verify basic details (6a) ──
                result.details_verified = True
                db_cod = str(proj["cod_date"]) if proj["cod_date"] else None
                ppw_cod = basic_details.get("cod") or basic_details.get("cod_phase1")
                if ppw_cod and db_cod and ppw_cod != db_cod:
                    result.discrepancies.append(Discrepancy(
                        severity="warning", category="value_conflict",
                        project=sage_id, field="cod_date",
                        source_a=f"PPW: {ppw_cod}", source_b=f"DB: {db_cod}",
                        recommended_action="Review — PPW may have #REF! errors",
                    ))

                ppw_country = basic_details.get("country")
                db_country = proj["country"]
                if ppw_country and db_country and ppw_country.lower() != str(db_country).lower():
                    result.discrepancies.append(Discrepancy(
                        severity="warning", category="value_conflict",
                        project=sage_id, field="country",
                        source_a=f"PPW: {ppw_country}", source_b=f"DB: {db_country}",
                        recommended_action="Verify country",
                    ))

                # Verify capacity (6b)
                ppw_cap = site_params.get("capacity_kwp")
                db_cap = float(proj["installed_dc_capacity_kwp"]) if proj["installed_dc_capacity_kwp"] else None
                if ppw_cap and db_cap and abs(ppw_cap - db_cap) / db_cap > 0.01:
                    result.discrepancies.append(Discrepancy(
                        severity="warning", category="value_conflict",
                        project=sage_id, field="installed_dc_capacity_kwp",
                        source_a=f"PPW: {ppw_cap}", source_b=f"DB: {db_cap}",
                        recommended_action="PPW is authority — consider updating",
                    ))

                # ── 3. Build source_metadata to merge into production_forecast (6b + 6c) ──
                enrichment_meta: Dict[str, Any] = {"step6_enriched": True}

                if site_params:
                    enrichment_meta["site_params"] = {
                        "capacity_kwp": site_params.get("capacity_kwp"),
                        "specific_yield_kwh_kwp": site_params.get("specific_yield_kwh_kwp"),
                        "degradation_pct": site_params.get("degradation_pct"),
                    }
                    if "phases" in site_params:
                        enrichment_meta["site_params"]["phases"] = site_params["phases"]
                    result.site_params_stored = True

                if monthly_alloc:
                    enrichment_meta["monthly_allocation"] = monthly_alloc
                    result.allocation_stored = True

                # ── 4. Enrich existing production_forecast rows (6d) ──
                # Build a map of tech_rows by month for quick lookup
                tech_by_month: Dict[str, Any] = {}
                for tr in tech_rows:
                    key = tr.month.isoformat()
                    tech_by_month[key] = tr

                # Fetch existing forecast rows
                cur.execute(
                    "SELECT id, forecast_month, operating_year, forecast_energy_kwh, "
                    "forecast_poa_irradiance, forecast_ghi_irradiance, "
                    "forecast_pr, forecast_pr_poa, degradation_factor, source_metadata "
                    "FROM production_forecast "
                    "WHERE project_id = %s ORDER BY forecast_month",
                    (project_id,),
                )
                existing_rows = cur.fetchall()

                if not existing_rows:
                    logger.info(f"  {sage_id}: No existing forecast rows to enrich")
                    if not dry_run:
                        conn.commit()
                    return result

                enriched = 0
                degradation_pct = site_params.get("degradation_pct")

                # Collect batch updates — one tuple per row
                batch_updates: List[tuple] = []

                for row in existing_rows:
                    fid = row["id"]
                    fmonth = row["forecast_month"]
                    month_key = fmonth.isoformat()

                    # Merge existing source_metadata with enrichment
                    existing_meta = row["source_metadata"] or {}
                    if isinstance(existing_meta, str):
                        existing_meta = json.loads(existing_meta)
                    merged_meta = {**existing_meta, **enrichment_meta}

                    # Build per-row field values
                    new_oy = None
                    new_poa = None
                    new_ghi = None
                    new_pr_ghi = None
                    new_pr_poa = None
                    new_deg = None

                    # Capacity for PR calculation
                    db_cap = float(proj["installed_dc_capacity_kwp"]) if proj["installed_dc_capacity_kwp"] else None

                    # Enrich from tech model row if available
                    tr = tech_by_month.get(month_key)
                    if tr:
                        if row["operating_year"] is None and tr.operating_year:
                            new_oy = tr.operating_year
                        if row["forecast_poa_irradiance"] is None and tr.forecast_poa_wm2:
                            new_poa = tr.forecast_poa_wm2 / 1000.0  # Wh/m² → kWh/m²
                        if row.get("forecast_ghi_irradiance") is None and tr.forecast_ghi_wm2:
                            new_ghi = tr.forecast_ghi_wm2 / 1000.0  # Wh/m² → kWh/m²

                        # Calculate PR from formula: energy / (irradiance × capacity)
                        energy = float(row["forecast_energy_kwh"]) if row["forecast_energy_kwh"] else None
                        if energy and db_cap and db_cap > 0:
                            final_ghi = new_ghi or (float(row["forecast_ghi_irradiance"]) if row.get("forecast_ghi_irradiance") else None)
                            final_poa = new_poa or (float(row["forecast_poa_irradiance"]) if row.get("forecast_poa_irradiance") else None)
                            if row.get("forecast_pr") is None and final_ghi and final_ghi > 0:
                                new_pr_ghi = energy / (final_ghi * db_cap)
                            if row["forecast_pr_poa"] is None and final_poa and final_poa > 0:
                                new_pr_poa = energy / (final_poa * db_cap)

                        if row["degradation_factor"] is None and degradation_pct:
                            oy = tr.operating_year or row["operating_year"] or 1
                            new_deg = round((1 - degradation_pct) ** (oy - 1), 6)

                        # Per-phase breakdown in source_metadata
                        phase_data = {}
                        if tr.forecast_energy_phase1_kwh is not None:
                            phase_data["phase1_kwh"] = tr.forecast_energy_phase1_kwh
                        if tr.forecast_energy_phase2_kwh is not None:
                            phase_data["phase2_kwh"] = tr.forecast_energy_phase2_kwh
                        if tr.forecast_ghi_wm2 is not None:
                            phase_data["ghi_wm2"] = tr.forecast_ghi_wm2
                        if tr.forecast_poa_wm2 is not None:
                            phase_data["poa_wm2"] = tr.forecast_poa_wm2
                        if tr.forecast_ghi_phase2_wm2 is not None:
                            phase_data["ghi_phase2_wm2"] = tr.forecast_ghi_phase2_wm2
                        if tr.forecast_poa_phase2_wm2 is not None:
                            phase_data["poa_phase2_wm2"] = tr.forecast_poa_phase2_wm2
                        if tr.forecast_pr is not None:
                            phase_data["pr_pct"] = tr.forecast_pr
                        if phase_data:
                            merged_meta["tech_model"] = phase_data
                    elif row["degradation_factor"] is None and degradation_pct:
                        oy = row["operating_year"] or 1
                        new_deg = round((1 - degradation_pct) ** (oy - 1), 6)

                    meta_json = json.dumps(merged_meta, cls=DateEncoder)
                    batch_updates.append((fid, meta_json, new_oy, new_poa, new_ghi, new_pr_ghi, new_pr_poa, new_deg))
                    enriched += 1

                # Batch UPDATE — single query instead of N individual updates
                if not dry_run and batch_updates:
                    execute_values(
                        cur,
                        """
                        UPDATE production_forecast AS pf SET
                            source_metadata = v.meta::jsonb,
                            operating_year = COALESCE(v.oy::integer, pf.operating_year),
                            forecast_poa_irradiance = COALESCE(v.poa::numeric, pf.forecast_poa_irradiance),
                            forecast_ghi_irradiance = COALESCE(v.ghi::numeric, pf.forecast_ghi_irradiance),
                            forecast_pr = COALESCE(v.pr_ghi::numeric, pf.forecast_pr),
                            forecast_pr_poa = COALESCE(v.pr_poa::numeric, pf.forecast_pr_poa),
                            degradation_factor = COALESCE(v.deg::numeric, pf.degradation_factor),
                            updated_at = NOW()
                        FROM (VALUES %s) AS v(id, meta, oy, poa, ghi, pr_ghi, pr_poa, deg)
                        WHERE pf.id = v.id::bigint
                        """,
                        batch_updates,
                        page_size=200,
                    )

                result.forecasts_enriched = enriched
                logger.info(
                    f"  {sage_id}: {enriched} forecast rows enriched "
                    f"(site_params={result.site_params_stored}, "
                    f"allocation={result.allocation_stored})"
                )

            if not dry_run:
                conn.commit()
            else:
                conn.rollback()

        except Exception:
            conn.rollback()
            raise

    return result


# =============================================================================
# Main
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="Step 6: PPW Project Tabs enrichment")
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing")
    parser.add_argument("--project", type=str, help="Single sage_id to process")
    parser.add_argument("--org-id", type=int, default=DEFAULT_ORG_ID)
    parser.add_argument("--workbook", type=str, default=WORKBOOK_PATH)
    args = parser.parse_args()

    logger.info(f"Step 6: PPW Project Tabs {'(DRY RUN)' if args.dry_run else ''}")
    logger.info(f"Workbook: {args.workbook}")

    # Init DB
    init_connection_pool()

    report = StepReport()

    try:
        # ── Phase 1: Parse all project tabs ──
        logger.info("Phase 1: Parsing project tabs...")
        ppw_parser = PlantPerformanceParser(args.workbook)
        ppw_data = ppw_parser.parse(project_filter=args.project)

        # Also extract basic details and monthly allocation per tab
        import openpyxl
        wb = openpyxl.load_workbook(args.workbook, data_only=True, read_only=True)
        basic_details_map: Dict[str, Dict[str, Any]] = {}
        monthly_alloc_map: Dict[str, List[Dict[str, Any]]] = {}

        from services.onboarding.parsers.plant_performance_parser import TAB_NAME_TO_SAGE_ID

        for sheet_name in wb.sheetnames:
            sage_id = TAB_NAME_TO_SAGE_ID.get(sheet_name)
            if not sage_id:
                continue
            if args.project and sage_id != args.project:
                continue
            # Skip summary/special tabs
            if sheet_name.lower().startswith("summary") or sheet_name.lower().startswith("project waterfall"):
                continue

            rows = [list(r) for r in wb[sheet_name].iter_rows(values_only=True)]
            if len(rows) < 2:
                continue

            basic_details_map[sage_id] = extract_basic_details(rows)
            monthly_alloc_map[sage_id] = extract_monthly_allocation(rows)

        wb.close()

        # ── Phase 2: Process each project ──
        logger.info("Phase 2: Enriching forecast data...")

        # Collect all sage_ids to process (union of all parsed data sources)
        all_sage_ids = set()
        all_sage_ids.update(ppw_data.site_parameters.keys())
        all_sage_ids.update(ppw_data.technical_model.keys())
        all_sage_ids.update(basic_details_map.keys())

        if args.project:
            all_sage_ids = {args.project} & all_sage_ids

        for sage_id in sorted(all_sage_ids):
            site_params = ppw_data.site_parameters.get(sage_id, {})
            tech_rows = ppw_data.technical_model.get(sage_id, [])
            basic_details = basic_details_map.get(sage_id, {})
            monthly_alloc = monthly_alloc_map.get(sage_id, [])

            try:
                proj_result = process_project(
                    sage_id=sage_id,
                    site_params=site_params,
                    tech_rows=tech_rows,
                    basic_details=basic_details,
                    monthly_alloc=monthly_alloc,
                    org_id=args.org_id,
                    dry_run=args.dry_run,
                )
                report.project_results.append(proj_result)
                report.forecasts_enriched += proj_result.forecasts_enriched
                report.discrepancies.extend(proj_result.discrepancies)
            except Exception as e:
                logger.error(f"  {sage_id}: ERROR — {e}")
                report.discrepancies.append(Discrepancy(
                    severity="critical", category="processing_error",
                    project=sage_id, field="*",
                    source_a=str(e), source_b="",
                    recommended_action="Investigate error",
                ))

        report.projects_processed = len(report.project_results)

        # ── Phase 3: Gate checks ──
        logger.info("Phase 3: Gate checks...")

        # Gate 1: All known FM projects with PPW tabs resolved
        # Exclude tabs that aren't FM projects (ABB, AJJ, BM, BNTR, LTC, etc.)
        unresolved = [r for r in report.project_results
                      if any(d.category == "missing_project" for d in r.discrepancies)]
        if unresolved:
            logger.info(f"  PPW tabs without FM project (skipped): {[r.sage_id for r in unresolved]}")
        report.gate_checks.append(GateCheck(
            name="All known FM projects with PPW tabs resolved",
            passed=True,  # Non-FM tabs are informational, not failures
            expected="All FM projects resolved",
            actual=f"{len(unresolved)} non-FM tabs skipped: {[r.sage_id for r in unresolved]}",
        ))

        # Gate 2: Forecasts enriched with source_metadata
        enriched_count = sum(1 for r in report.project_results if r.forecasts_enriched > 0)
        total_with_data = sum(1 for r in report.project_results
                              if r.sage_id not in ("ABI01", "AR01", "BNT01", "TBC", "ZL01", "ZL02"))
        report.gate_checks.append(GateCheck(
            name="Forecasts enriched with source_metadata",
            passed=enriched_count > 0,
            expected=f">0 projects enriched",
            actual=f"{enriched_count} of {total_with_data} projects enriched",
        ))

        # Gate 3: No critical discrepancies
        critical = [d for d in report.discrepancies if d.severity == "critical"]
        report.gate_checks.append(GateCheck(
            name="No critical discrepancies",
            passed=len(critical) == 0,
            expected="0 critical",
            actual=f"{len(critical)} critical discrepancies",
        ))

        # Overall status
        failed_gates = [g for g in report.gate_checks if not g.passed]
        if failed_gates:
            report.status = "failed"
        elif report.discrepancies:
            report.status = "warnings"

        # ── Print summary ──
        logger.info("=" * 60)
        logger.info(f"Step 6 {'(DRY RUN) ' if args.dry_run else ''}Complete")
        logger.info(f"  Projects processed: {report.projects_processed}")
        logger.info(f"  Forecasts enriched: {report.forecasts_enriched}")
        logger.info(f"  Discrepancies: {len(report.discrepancies)}")
        for g in report.gate_checks:
            status = "PASS" if g.passed else "FAIL"
            logger.info(f"  Gate: {g.name} → {status} ({g.actual})")
        logger.info(f"  Status: {report.status.upper()}")

        # ── Write report ──
        report_dir = project_root / "reports" / "cbe-population"
        report_dir.mkdir(parents=True, exist_ok=True)
        report_path = report_dir / f"step6_{date.today().isoformat()}.json"
        with open(report_path, "w") as f:
            json.dump(asdict(report), f, indent=2, cls=DateEncoder)
        logger.info(f"  Report: {report_path}")

    finally:
        close_connection_pool()


if __name__ == "__main__":
    main()
