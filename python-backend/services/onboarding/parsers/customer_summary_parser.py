"""
Parser for CBE Customer summary.xlsx.

Extracts project metadata, counterparty info, and contract type from the
Customer Summary workbook (~45 rows, cols A-N). Groups multi-row projects
(e.g., LOI01 Solar+BESS, QMM01 Solar+Wind+BESS) by sage_customer_id.
"""

import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, List, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass
class CustomerSummaryRow:
    row_number: int
    cbe_entity: Optional[str] = None
    sage_company: Optional[str] = None
    external_project_id: Optional[str] = None
    contract_type: Optional[str] = None
    revenue_subtype: Optional[str] = None
    tenor_years: Optional[int] = None
    cod_date: Optional[date] = None
    size_kwp: Optional[float] = None
    project_name: Optional[str] = None
    customer_name: Optional[str] = None
    industry: Optional[str] = None
    country: Optional[str] = None
    sage_customer_id: Optional[str] = None
    billing_currency: Optional[str] = None


@dataclass
class CustomerSummaryProject:
    """Deduplicated project from one or more Customer Summary rows."""
    sage_id: str
    project_name: str
    customer_name: Optional[str] = None
    cbe_entity: Optional[str] = None
    sage_company: Optional[str] = None
    external_project_id: Optional[str] = None
    contract_type: Optional[str] = None
    revenue_subtype: Optional[str] = None
    tenor_years: Optional[int] = None
    cod_date: Optional[date] = None
    size_kwp: Optional[float] = None
    industry: Optional[str] = None
    country: Optional[str] = None
    billing_currency: Optional[str] = None
    sub_rows: List[CustomerSummaryRow] = field(default_factory=list)


# XFlora special case: col M contains "XF-AB/BV/L01/SS" -> sage_id = "XFAB"
# Only the primary (XFAB) appears in the Customer summary xlsx
XFLORA_PREFIX = "XF-"


class CustomerSummaryParser:
    """Parser for CBE Customer summary.xlsx."""

    def parse(self, file_path: str) -> List[CustomerSummaryRow]:
        """
        Parse Customer summary.xlsx into raw rows.

        Args:
            file_path: Path to the xlsx file.

        Returns:
            List of parsed rows (empty/separator rows skipped).
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Customer summary file not found: {file_path}")

        logger.info(f"Parsing Customer summary: {file_path}")
        wb = load_workbook(str(path), data_only=True, read_only=True)
        ws = wb.active
        rows: List[CustomerSummaryRow] = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            cells = [c.value for c in row]

            # Skip empty separator rows
            if not any(cells):
                continue

            # Need at least sage_customer_id (col M) or project_name (col I) to be useful
            sage_customer_id = self._clean_str(cells[12]) if len(cells) > 12 else None
            project_name = self._clean_str(cells[8]) if len(cells) > 8 else None
            if not sage_customer_id and not project_name:
                continue

            parsed = CustomerSummaryRow(
                row_number=row_idx,
                cbe_entity=self._clean_str(cells[0]) if len(cells) > 0 else None,
                sage_company=self._clean_str(cells[1]) if len(cells) > 1 else None,
                external_project_id=self._clean_str(cells[2]) if len(cells) > 2 else None,
                contract_type=self._clean_str(cells[3]) if len(cells) > 3 else None,
                revenue_subtype=self._clean_str(cells[4]) if len(cells) > 4 else None,
                tenor_years=self._to_int(cells[5]) if len(cells) > 5 else None,
                cod_date=self._to_date(cells[6]) if len(cells) > 6 else None,
                size_kwp=self._to_float(cells[7]) if len(cells) > 7 else None,
                project_name=project_name,
                customer_name=self._clean_str(cells[9]) if len(cells) > 9 else None,
                industry=self._clean_str(cells[10]) if len(cells) > 10 else None,
                country=self._clean_str(cells[11]) if len(cells) > 11 else None,
                sage_customer_id=sage_customer_id,
                billing_currency=self._normalize_currency(cells[13]) if len(cells) > 13 else None,
            )
            rows.append(parsed)

        wb.close()
        logger.info(f"Parsed {len(rows)} rows from Customer summary")
        return rows

    def deduplicate_to_projects(
        self, rows: List[CustomerSummaryRow]
    ) -> List[CustomerSummaryProject]:
        """
        Group rows by sage_customer_id to produce deduplicated projects.

        Multi-row projects (e.g., LOI01 with Solar+BESS sub-components)
        are merged: primary row fields come from the first row, sub_rows
        captures all rows for that sage_id.
        """
        groups: dict[str, List[CustomerSummaryRow]] = {}

        for row in rows:
            sage_id = self._resolve_sage_id(row.sage_customer_id)
            if not sage_id:
                logger.warning(
                    f"Row {row.row_number}: no sage_customer_id, skipping"
                )
                continue

            groups.setdefault(sage_id, []).append(row)

        projects: List[CustomerSummaryProject] = []
        for sage_id, group_rows in groups.items():
            primary = group_rows[0]

            # For multi-row projects, sum size_kwp across sub-rows
            total_kwp = None
            kwp_values = [r.size_kwp for r in group_rows if r.size_kwp is not None]
            if kwp_values:
                total_kwp = sum(kwp_values)

            project = CustomerSummaryProject(
                sage_id=sage_id,
                project_name=primary.project_name or sage_id,
                customer_name=primary.customer_name,
                cbe_entity=primary.cbe_entity,
                sage_company=primary.sage_company,
                external_project_id=primary.external_project_id,
                contract_type=primary.contract_type,
                revenue_subtype=primary.revenue_subtype,
                tenor_years=primary.tenor_years,
                cod_date=primary.cod_date,
                size_kwp=total_kwp,
                industry=primary.industry,
                country=primary.country,
                billing_currency=primary.billing_currency,
                sub_rows=group_rows,
            )
            projects.append(project)

        logger.info(
            f"Deduplicated {len(rows)} rows into {len(projects)} projects"
        )
        return projects

    # ─── Helpers ──────────────────────────────────────────────────────────────

    @staticmethod
    def _resolve_sage_id(raw: Optional[str]) -> Optional[str]:
        """Resolve raw sage_customer_id to canonical sage_id."""
        if not raw:
            return None
        cleaned = raw.strip()
        # XFlora special case: "XF-AB/BV/L01/SS" -> "XFAB"
        # Only the primary project (XFAB) appears in Customer summary
        if cleaned.startswith(XFLORA_PREFIX):
            return "XFAB"
        return cleaned

    @staticmethod
    def _clean_str(value: Any) -> Optional[str]:
        """Strip whitespace from string values."""
        if value is None:
            return None
        s = str(value).strip()
        return s if s else None

    @staticmethod
    def _normalize_currency(value: Any) -> Optional[str]:
        """Normalize currency codes (e.g., 'kES' -> 'KES')."""
        if value is None:
            return None
        s = str(value).strip().upper()
        return s if s else None

    @staticmethod
    def _to_date(value: Any) -> Optional[date]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _to_float(value: Any) -> Optional[float]:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        try:
            cleaned = re.sub(r'[,\s]', '', str(value))
            cleaned = re.sub(r'(kwp|kw|mw)$', '', cleaned, flags=re.IGNORECASE)
            return float(cleaned)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _to_int(value: Any) -> Optional[int]:
        if value is None:
            return None
        if isinstance(value, int):
            return value
        try:
            return int(float(str(value).replace(",", "").strip()))
        except (ValueError, TypeError):
            return None
