"""
Market Reference Pricing parser (.xlsx) for cross-examination pipeline.

Parses: 'Sage Contract Extracts market Ref pricing data.xlsx'
Tabs:
  - PO Summary: per-project tariff summary
  - Per-project sheets (KAS01, MOH001, etc.): monthly MRP observations
  - Market reference pricing sheets: historical reference prices
"""

import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from models.onboarding import MarketRefPricingData, MarketRefPricingProject

logger = logging.getLogger(__name__)

# ─── Known Column Headers for PO Summary ───────────────────────────────────
PO_SUMMARY_HEADERS = {
    "project", "customer", "site", "sage_id", "sage id", "facility",
    "base rate", "tariff", "discount", "floor", "ceiling",
    "currency", "escalation", "type", "formula",
}

# ─── Sheet name → sage_id mapping (handles naming mismatches) ──────────────
SHEET_TO_SAGE_ID = {
    "MOH001": "MOH01",
}

# ─── Known ZPRICODE labels → component names ──────────────────────────────
COMPONENT_LABEL_MAP = {
    "energy charge": "energy_charge",
    "govt levy": "govt_levy",
    "government levy": "govt_levy",
    "street lighting": "street_lighting",
    "subsidy": "subsidy",
    "consumption": "consumption",
    "erc": "erc",
    "fcc": "fcc",
    "ferfa": "ferfa",
    "ia": "ia",
    "rep": "rep",
    "warma": "warma",
    "grid reference price": "grid_reference_price",
    "diesel fuel price": "diesel_fuel_price",
}


def _safe_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).strip().replace(",", "").replace("%", "")
        if s in ("", "-", "N/A", "n/a", "none"):
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def _detect_sage_id(name: str) -> Optional[str]:
    """Try to extract sage_id from project/customer name."""
    if not name:
        return None
    parts = name.strip().split()
    if parts:
        candidate = parts[0].strip(" -–—")
        if len(candidate) >= 3 and any(c.isdigit() for c in candidate):
            return candidate.upper()
    return None


class MarketRefPricingParser:
    """
    Parse the Market Reference Pricing workbook (.xlsx).

    Usage:
        parser = MarketRefPricingParser("/path/to/market_ref.xlsx")
        data = parser.parse()
        kas01 = data.projects.get("KAS01")
    """

    def __init__(self, file_path: str):
        self.file_path = file_path

    def parse(self, project_filter: Optional[str] = None) -> MarketRefPricingData:
        """Parse the workbook and return structured pricing data."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for .xlsx file support")

        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Market Ref Pricing file not found: {self.file_path}")

        logger.info(f"Parsing Market Ref Pricing: {self.file_path}")

        wb = openpyxl.load_workbook(self.file_path, data_only=True, read_only=True)
        result = MarketRefPricingData()

        # Parse PO Summary tab
        po_sheet = self._find_sheet(wb.sheetnames, ["PO Summary", "PO_Summary", "Summary", "PO"])
        if po_sheet:
            result.po_summary, projects = self._parse_po_summary(wb[po_sheet], project_filter)
            result.projects.update(projects)
            logger.info(f"Parsed {len(projects)} projects from PO Summary")
        else:
            logger.warning("PO Summary sheet not found")

        # Parse market reference pricing sheets
        for sheet_name in wb.sheetnames:
            sn_lower = sheet_name.lower()
            if any(kw in sn_lower for kw in ["reference", "market", "grid price", "gen price", "tariff"]):
                if sheet_name == po_sheet:
                    continue
                ref_prices = self._parse_reference_prices(wb[sheet_name], project_filter)
                for sage_id, prices in ref_prices.items():
                    if sage_id in result.projects:
                        result.projects[sage_id].reference_prices.update(prices)
                    else:
                        result.projects[sage_id] = MarketRefPricingProject(
                            sage_id=sage_id,
                            reference_prices=prices,
                        )

        wb.close()
        logger.info(f"Parsed market ref pricing for {len(result.projects)} project(s)")
        return result

    def _find_sheet(self, sheet_names: List[str], candidates: List[str]) -> Optional[str]:
        """Find sheet by candidate names."""
        lower_map = {s.lower(): s for s in sheet_names}
        for c in candidates:
            if c.lower() in lower_map:
                return lower_map[c.lower()]
        for c in candidates:
            for s in sheet_names:
                if c.lower() in s.lower():
                    return s
        return None

    def get_available_sheets(self) -> List[str]:
        """Return sage_ids for all project sheets in the workbook (excluding Summary)."""
        import openpyxl

        if not os.path.exists(self.file_path):
            return []
        wb = openpyxl.load_workbook(self.file_path, read_only=True)
        sheets = []
        summary_sheet = self._find_sheet(wb.sheetnames, ["PO Summary", "PO_Summary", "Summary", "PO"])
        for name in wb.sheetnames:
            if name == summary_sheet:
                continue
            sage_id = SHEET_TO_SAGE_ID.get(name, name).upper()
            sheets.append(sage_id)
        wb.close()
        return sheets

    def parse_mrp_monthly(self, sage_id: str) -> List[Dict]:
        """
        Parse monthly MRP observations from a project-specific sheet.

        Handles varied sheet layouts:
        - Standard: ZDAT, ZPRICODE1..N, ZPRITOT (KAS01, UGL01, GBL01)
        - Multi-section: multiple MRP sections side-by-side (MOH001) — uses first section
        - No components: just ZDAT + ZPRITOT (NBL01, NBL02)
        - Different components: Kenyan tariff codes (UTK01, TBM01)
        - Offset headers: ZDAT not in first column (JAB01)

        Returns list of:
        {
            "billing_month": "2021-01",       # YYYY-MM
            "mrp_per_kwh": 0.835,             # ZPRITOT / Price total
            "tariff_components": {            # raw component charges (if available)
                "energy_charge": 0.7952,
                "govt_levy": 0.0159,
                "street_lighting": 0.0239,
            }
        }
        """
        import openpyxl

        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Market Ref Pricing file not found: {self.file_path}")

        wb = openpyxl.load_workbook(self.file_path, data_only=True, read_only=True)

        # Find sheet: try sage_id directly, then reverse-map (e.g. MOH01 → MOH001)
        reverse_map = {v: k for k, v in SHEET_TO_SAGE_ID.items()}
        sheet_name = None
        for candidate in [sage_id, reverse_map.get(sage_id, "")]:
            if not candidate:
                continue
            for sn in wb.sheetnames:
                if sn.upper() == candidate.upper():
                    sheet_name = sn
                    break
            if sheet_name:
                break

        if not sheet_name:
            wb.close()
            raise ValueError(f"No sheet found for sage_id={sage_id} in {self.file_path}")

        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        wb.close()

        if len(rows) < 3:
            return []

        # ── Locate the header rows ──
        # Two patterns:
        #   A) Code row with ZDAT (e.g. "ZDAT, ZPRICODE1, ..., ZPRITOT") + label row below
        #   B) Label-only row with "Period" (e.g. "Period, Consumption, ..., Price total")
        # For multi-section sheets (GRID + GENERATOR side-by-side), we use the first
        # ZDAT/Period block only — stop scanning columns at the first None gap after ZDAT.
        zdat_row_idx = None
        zdat_col = None
        has_code_row = False

        for i, row in enumerate(rows[:20]):
            for j, val in enumerate(row):
                if val is None:
                    continue
                v = str(val).strip().upper()
                if v == "ZDAT":
                    zdat_row_idx = i
                    zdat_col = j
                    has_code_row = True
                    break
            if zdat_row_idx is not None:
                break

        if zdat_row_idx is None:
            # Fallback: look for "Period" as header (no code row above)
            for i, row in enumerate(rows[:20]):
                for j, val in enumerate(row):
                    if val is None:
                        continue
                    v = str(val).strip().lower()
                    if v == "period":
                        zdat_row_idx = i
                        zdat_col = j
                        has_code_row = False
                        break
                if zdat_row_idx is not None:
                    break

        if zdat_row_idx is None:
            logger.warning(f"No ZDAT/Period header found in sheet {sheet_name}")
            return []

        # ── Determine the column range for the first (grid) section ──
        # Stop at the first None column after zdat_col to avoid picking up generator section
        section_end = len(rows[zdat_row_idx])
        for j in range(zdat_col + 1, len(rows[zdat_row_idx])):
            if rows[zdat_row_idx][j] is None:
                section_end = j
                break

        # ── Map columns from the code row (if present) ──
        zpritot_col = None
        zpricode_cols: Dict[int, str] = {}

        if has_code_row:
            code_row = rows[zdat_row_idx]
            for j in range(zdat_col + 1, section_end):
                val = code_row[j] if j < len(code_row) else None
                if val is None:
                    continue
                v = str(val).strip().upper()
                if v == "ZPRITOT" and zpritot_col is None:
                    zpritot_col = j
                elif v.startswith("ZPRICODE"):
                    zpricode_cols[j] = v

        # ── Map human-readable labels ──
        # If has_code_row: label row is zdat_row_idx + 1
        # If no code row: the zdat_row IS the label row
        label_row_idx = (zdat_row_idx + 1) if has_code_row else zdat_row_idx
        component_cols: Dict[int, str] = {}

        if label_row_idx < len(rows):
            label_row = rows[label_row_idx]
            for j in range(zdat_col + 1, section_end):
                val = label_row[j] if j < len(label_row) else None
                if val is None:
                    continue
                label = str(val).strip()
                label_lower = label.lower()
                # Strip parenthetical currency, e.g. "Energy Charge(GHS)" -> "energy charge"
                clean = label_lower.split("(")[0].strip()
                if clean == "price total":
                    if zpritot_col is None:
                        zpritot_col = j
                    continue
                if clean == "period":
                    continue
                # Map to known component name, or use sanitized label
                mapped = COMPONENT_LABEL_MAP.get(clean)
                if mapped:
                    component_cols[j] = mapped
                elif j in zpricode_cols:
                    component_cols[j] = clean.replace(" ", "_").replace("(", "").replace(")", "")
                elif not has_code_row:
                    # No code row — accept any recognized label as a component
                    if mapped:
                        component_cols[j] = mapped

        if zpritot_col is None:
            logger.warning(f"No ZPRITOT/Price total column found in sheet {sheet_name}")
            return []

        # ── Data starts after the label row ──
        data_start = label_row_idx + 1

        observations: List[Dict] = []
        for i in range(data_start, len(rows)):
            row = rows[i]
            if not row or all(v is None for v in row):
                continue

            # Get date value
            if zdat_col >= len(row):
                continue
            date_val = row[zdat_col]
            if date_val is None:
                continue

            # Validate it's a date
            if isinstance(date_val, datetime):
                period_date = date_val
            elif isinstance(date_val, str):
                # Try parsing common date formats
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                    try:
                        period_date = datetime.strptime(date_val.strip(), fmt)
                        break
                    except ValueError:
                        continue
                else:
                    # Not a date — sentinel row (e.g. ZPRIAVG), stop
                    break
            else:
                break

            # Get MRP (ZPRITOT)
            mrp_val = _safe_float(row[zpritot_col] if zpritot_col < len(row) else None)
            if mrp_val is None:
                continue

            # Get component breakdown
            components = {}
            for col_idx, comp_name in component_cols.items():
                if col_idx < len(row):
                    comp_val = _safe_float(row[col_idx])
                    if comp_val is not None:
                        components[comp_name] = comp_val

            billing_month = period_date.strftime("%Y-%m")
            observations.append({
                "billing_month": billing_month,
                "mrp_per_kwh": mrp_val,
                "tariff_components": components,
            })

        logger.info(f"Parsed {len(observations)} monthly MRP observations from sheet {sheet_name}")
        return observations

    def _parse_po_summary(
        self, ws: Any, project_filter: Optional[str] = None
    ) -> Tuple[List[Dict], Dict[str, MarketRefPricingProject]]:
        """Parse PO Summary tab for per-project tariff summary."""
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))

        if len(rows) < 2:
            return [], {}

        # Find header row
        header_idx = None
        col_map: Dict[str, int] = {}

        for i, row in enumerate(rows[:20]):
            if not row:
                continue
            matches = 0
            for j, val in enumerate(row):
                if val is None:
                    continue
                v = str(val).strip().lower()
                if v in PO_SUMMARY_HEADERS:
                    col_map[v] = j
                    matches += 1
            if matches >= 3:
                header_idx = i
                break

        if header_idx is None:
            return [], {}

        # Resolve column indices to field names
        sage_col = col_map.get("sage_id") or col_map.get("sage id") or col_map.get("facility")
        project_col = col_map.get("project") or col_map.get("customer") or col_map.get("site")
        base_rate_col = col_map.get("base rate") or col_map.get("tariff")
        discount_col = col_map.get("discount")
        floor_col = col_map.get("floor")
        ceiling_col = col_map.get("ceiling")
        currency_col = col_map.get("currency")

        po_rows: List[Dict] = []
        projects: Dict[str, MarketRefPricingProject] = {}

        for i in range(header_idx + 1, len(rows)):
            row = rows[i]
            if not row or all(v is None for v in row):
                continue

            # Determine sage_id
            sage_id = None
            if sage_col is not None and sage_col < len(row) and row[sage_col]:
                sage_id = str(row[sage_col]).strip().upper()
            if not sage_id and project_col is not None and project_col < len(row):
                sage_id = _detect_sage_id(str(row[project_col] or ""))
            if not sage_id:
                continue

            if project_filter and sage_id != project_filter:
                continue

            project_name = str(row[project_col]).strip() if project_col is not None and project_col < len(row) and row[project_col] else None

            tariff_summary = {}
            if base_rate_col is not None and base_rate_col < len(row):
                tariff_summary["base_rate"] = _safe_float(row[base_rate_col])
            if discount_col is not None and discount_col < len(row):
                disc = _safe_float(row[discount_col])
                if disc is not None and disc > 1.0:
                    disc = disc / 100.0
                tariff_summary["discount_pct"] = disc
            if floor_col is not None and floor_col < len(row):
                tariff_summary["floor_rate"] = _safe_float(row[floor_col])
            if ceiling_col is not None and ceiling_col < len(row):
                tariff_summary["ceiling_rate"] = _safe_float(row[ceiling_col])
            if currency_col is not None and currency_col < len(row):
                tariff_summary["currency"] = str(row[currency_col] or "").strip().upper()

            po_row_dict = {"sage_id": sage_id, "project_name": project_name, **tariff_summary}
            po_rows.append(po_row_dict)

            projects[sage_id] = MarketRefPricingProject(
                sage_id=sage_id,
                project_name=project_name,
                tariff_summary=tariff_summary,
            )

        return po_rows, projects

    def _parse_reference_prices(
        self, ws: Any, project_filter: Optional[str] = None
    ) -> Dict[str, Dict[str, float]]:
        """Parse a reference pricing sheet for sage_id -> {period: price}."""
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))

        if len(rows) < 2:
            return {}

        # Find header with date/period columns
        header_idx = None
        period_cols: Dict[int, str] = {}

        for i, row in enumerate(rows[:20]):
            if not row:
                continue
            found = 0
            for j, val in enumerate(row):
                if val is None:
                    continue
                s = str(val).strip()
                # Look for date-like headers (years, month-year, etc.)
                try:
                    yr = int(s)
                    if 2015 <= yr <= 2050:
                        period_cols[j] = s
                        found += 1
                except ValueError:
                    pass
            if found >= 3:
                header_idx = i
                break

        if header_idx is None:
            return {}

        result: Dict[str, Dict[str, float]] = {}
        for i in range(header_idx + 1, len(rows)):
            row = rows[i]
            if not row or all(v is None for v in row):
                continue

            sage_id = None
            for val in row:
                if val is not None:
                    sage_id = _detect_sage_id(str(val)) or str(val).strip().upper()
                    break
            if not sage_id:
                continue
            if project_filter and sage_id != project_filter:
                continue

            prices: Dict[str, float] = {}
            for col_idx, period in period_cols.items():
                if col_idx < len(row):
                    price = _safe_float(row[col_idx])
                    if price is not None:
                        prices[period] = price

            if prices:
                result[sage_id] = prices

        return result
