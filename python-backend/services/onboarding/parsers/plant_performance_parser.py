"""
Plant Performance Workbook parser (.xlsx) for cross-examination pipeline.

Parses: 'Operations Plant Performance Workbook.xlsx' (51 tabs)
Extracts per tab: monthly metered energy, available energy, GHI irradiance,
POA irradiance, PR%, availability %.

For tabs with a Technical Model section (e.g. KAS01), extracts full
per-phase forecast + actual time-series including meter readings,
irradiance, PR, availability, and performance comparisons.

Maps tab names -> sage_ids via explicit lookup dictionary.
"""

import logging
import os
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from models.onboarding import (
    PlantPerformanceData,
    PlantPerformanceMonthly,
    SummaryPerformanceRow,
    TechnicalModelRow,
    WaterfallRow,
)

logger = logging.getLogger(__name__)

# ─── Tab Name → sage_id Mapping ────────────────────────────────────────────
# Explicit lookup (no fuzzy matching). Tab names in the workbook are
# project names, not sage_ids, so we map them explicitly.
TAB_NAME_TO_SAGE_ID: Dict[str, str] = {
    "Kasapreko": "KAS01",
    "KAS01": "KAS01",
    "Mohinani": "MOH01",
    "MOH01": "MOH01",
    "NBL": "NBL01",
    "NBL01": "NBL01",
    "NBL02": "NBL02",
    "Guinness": "GC001",
    "GC001": "GC001",
    "Unilever": "UGL01",
    "UGL01": "UGL01",
    "Loisaba": "LOI01",
    "LOI01": "LOI01",
    "XFlora": "XFAB",
    "XFAB": "XFAB",
    "QMM": "QMM01",
    "QMM01": "QMM01",
    "Jabana": "JAB01",
    "JAB01": "JAB01",
    "GBL": "GBL01",
    "GBL01": "GBL01",
    "UNSOS": "UNSOS",
    "Caledonia": "CAL01",
    "CAL01": "CAL01",
    "TWG": "TWG01",
    "TWG01": "TWG01",
    "Mirambo": "MIR01",
    "MIR01": "MIR01",
    "XF": "XFAB",
    "ERG": "ERG",
    "UTK01": "UTK01",
    "AMP01": "AMP01",
    "IVL01": "IVL01",
    "TBM01": "TBM01",
    "MB01": "MB01",
    "MF01": "MF01",
    "MP01": "MP01",
    "MP02": "MP02",
    "NC02": "NC02",
    "NC03": "NC03",
    "MOH01": "MOH01",
    "XFBV": "XFBV",
    "XFL01": "XFL01",
    "XFSS": "XFSS",
    "LTC": "LTC",
    "BM": "BM",
    "AJJ": "AJJ",
    "ABB": "ABB",
    "BNTR": "BNTR",
}

# ─── Known Column Headers ──────────────────────────────────────────────────
MONTH_NAMES = [
    "jan", "feb", "mar", "apr", "may", "jun",
    "jul", "aug", "sep", "oct", "nov", "dec",
]

ENERGY_KEYWORDS = ["metered", "energy", "kwh", "mwh", "generation", "production"]
AVAILABLE_KEYWORDS = ["available", "deemed"]
GHI_KEYWORDS = ["ghi", "global horizontal"]
POA_KEYWORDS = ["poa", "plane of array", "tilted"]
PR_KEYWORDS = ["pr", "performance ratio"]
AVAIL_KEYWORDS = ["availability", "avail %", "avail%"]

# ─── Technical Model column header fragments ──────────────────────────────
# Maps header substring (lowered) -> TechnicalModelRow field name.
# Order matters: more specific patterns first to avoid ambiguous matches.
TECH_MODEL_COLUMNS: List[Tuple[str, str]] = [
    # Forecast energy per phase
    ("forecast energy phase 1", "forecast_energy_phase1_kwh"),
    ("forecast energy phase 2", "forecast_energy_phase2_kwh"),
    ("forcast combined", "forecast_energy_combined_kwh"),
    ("forecast combined", "forecast_energy_combined_kwh"),
    ("forecast energy", "forecast_energy_combined_kwh"),
    # Irradiance — "irrad" (actual) MUST come before "irr" (forecast) since
    # "ghi irr" is a substring of "ghi irrad" and would greedily match actual columns.
    # Phase-specific forecast patterns are safe (contain "phase" which actual columns lack).
    ("ghi irr phase 2", "forecast_ghi_phase2_wm2"),
    ("ghi irr phase 1", "forecast_ghi_wm2"),
    ("ghi irr phase", "forecast_ghi_wm2"),
    ("poa irr phase 2", "forecast_poa_phase2_wm2"),
    ("poa irr phase 1", "forecast_poa_wm2"),
    ("poa irr phase", "forecast_poa_wm2"),
    # Actual irradiance — before generic "irr" to prevent greedy match
    ("ghi irrad", "actual_ghi_wm2"),
    ("poa irrad", "actual_poa_wm2"),
    # Generic forecast irradiance (fallback — only matches "GHI Irr", not "GHI Irrad")
    ("ghi irr", "forecast_ghi_wm2"),
    ("poa irr", "forecast_poa_wm2"),
    # Forecast PR
    ("pr ghi", "forecast_pr"),
    ("pr poa", "forecast_pr_poa"),
    # Actual per-phase meter readings
    ("phase-1 invoiced", "phase1_invoiced_kwh"),
    ("phase 1 invoiced", "phase1_invoiced_kwh"),
    ("phase-2 invoiced", "phase2_invoiced_kwh"),
    ("phase 2 invoiced", "phase2_invoiced_kwh"),
    # Actual aggregated
    ("phase 1 and 2 metered", "total_metered_kwh"),
    ("phase 1+2 metered", "total_metered_kwh"),
    ("total metered energy", "total_metered_kwh"),
    ("metered energy", "total_metered_kwh"),
    ("available energy", "available_energy_kwh"),
    ("total energy", "total_energy_kwh"),
    # Actual PR / Availability
    ("pr %", "actual_pr"),
    ("availability %", "actual_availability_pct"),
    ("availability%", "actual_availability_pct"),
    # Comparisons
    ("energy comparison", "energy_comparison"),
    ("irr comparison", "irr_comparison"),
    ("pr comparison", "pr_comparison"),
    ("comment", "comments"),
]

# Site parameter label fragments (rows 2-6)
SITE_PARAM_LABELS: Dict[str, str] = {
    "installed capacity": "capacity_kwp",
    "capacity": "capacity_kwp",
    "kwp": "capacity_kwp",
    "degradation": "degradation_pct",
    "specific yield": "specific_yield_kwh_kwp",
}


def _safe_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if val != 0 else None
    try:
        s = str(val).strip().replace(",", "").replace("%", "")
        if s in ("", "-", "N/A", "n/a"):
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def _safe_float_allow_zero(val: Any) -> Optional[float]:
    """Like _safe_float but allows zero values (needed for meter readings)."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).strip().replace(",", "").replace("%", "")
        if s in ("", "-", "N/A", "n/a"):
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def _detect_column_type(header: str) -> Optional[str]:
    """Detect what data a column contains from its header."""
    h = header.lower().strip()
    for kw in AVAILABLE_KEYWORDS:
        if kw in h:
            return "available_energy"
    for kw in ENERGY_KEYWORDS:
        if kw in h:
            return "metered_energy"
    for kw in GHI_KEYWORDS:
        if kw in h:
            return "ghi"
    for kw in POA_KEYWORDS:
        if kw in h:
            return "poa"
    for kw in PR_KEYWORDS:
        if kw in h:
            return "pr"
    for kw in AVAIL_KEYWORDS:
        if kw in h:
            return "availability"
    return None


def _parse_date(val: Any) -> Optional[date]:
    """Parse a date value from a cell, returning first-of-month."""
    if isinstance(val, datetime):
        return val.date().replace(day=1)
    if isinstance(val, date):
        return val.replace(day=1)
    if val:
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%b-%y", "%b %Y", "%m/%Y"):
            try:
                return datetime.strptime(str(val).strip(), fmt).date().replace(day=1)
            except ValueError:
                continue
    return None


class PlantPerformanceParser:
    """
    Parse the Plant Performance Workbook (.xlsx).

    Usage:
        parser = PlantPerformanceParser("/path/to/workbook.xlsx")
        data = parser.parse()
        kas01_monthly = data.projects.get("KAS01")
        kas01_tech = data.technical_model.get("KAS01")
    """

    def __init__(self, file_path: str):
        self.file_path = file_path

    def parse(self, project_filter: Optional[str] = None) -> PlantPerformanceData:
        """Parse all tabs and return monthly performance data per project."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for .xlsx file support")

        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Plant Performance Workbook not found: {self.file_path}")

        logger.info(f"Parsing Plant Performance Workbook: {self.file_path}")

        wb = openpyxl.load_workbook(self.file_path, data_only=True, read_only=True)
        result = PlantPerformanceData()

        for sheet_name in wb.sheetnames:
            # Resolve sage_id
            sage_id = TAB_NAME_TO_SAGE_ID.get(sheet_name)
            if not sage_id:
                # Try partial match
                for tab_key, sid in TAB_NAME_TO_SAGE_ID.items():
                    if tab_key.lower() in sheet_name.lower():
                        sage_id = sid
                        break
            if not sage_id:
                logger.debug(f"Skipping unmapped tab: {sheet_name}")
                continue

            if project_filter and sage_id != project_filter:
                continue

            result.tab_to_sage_id[sheet_name] = sage_id

            # Load all rows
            rows: List[List[Any]] = []
            for row in wb[sheet_name].iter_rows(values_only=True):
                rows.append(list(row))

            if len(rows) < 2:
                continue

            # Try Technical Model extraction first
            tech_rows = self._parse_technical_model(rows, sheet_name)
            if tech_rows:
                result.technical_model.setdefault(sage_id, []).extend(tech_rows)
                logger.info(f"  {sheet_name}: {len(tech_rows)} Technical Model rows")

            # Extract site parameters (rows 2-6)
            site_params = self._extract_site_parameters(rows, sheet_name)
            if site_params:
                result.site_parameters[sage_id] = site_params

            # Also parse with keyword-based method (populates projects dict)
            monthly = self._parse_sheet_from_rows(rows, sheet_name)
            if monthly:
                result.projects.setdefault(sage_id, []).extend(monthly)

        wb.close()
        logger.info(
            f"Parsed performance data for {len(result.projects)} project(s), "
            f"{len(result.technical_model)} with Technical Model"
        )
        return result

    # ─── Technical Model Parsing ──────────────────────────────────────────

    def _find_technical_model_header(
        self, rows: List[List[Any]]
    ) -> Optional[Tuple[int, Dict[int, str]]]:
        """
        Find the Technical Model header row.

        Returns (header_row_index, {col_index: field_name}) or None.
        The header is identified by having a "Date" column, an "OY" column,
        and at least 2 recognizable Technical Model column headers.
        """
        for i, row in enumerate(rows):
            if not row:
                continue
            # Stringify all cells for matching — normalize Excel newlines
            cells = [str(c).strip().replace('\n', ' ').replace('\r', '') if c is not None else "" for c in row]
            cells_lower = [c.lower() for c in cells]

            # Must have "Date" and "OY" columns
            has_date = any(c in ("date", "month") for c in cells_lower)
            has_oy = "oy" in cells_lower
            if not has_date or not has_oy:
                continue

            # Map columns to field names
            col_map: Dict[int, str] = {}
            date_col: Optional[int] = None
            oy_col: Optional[int] = None

            for j, cell_lower in enumerate(cells_lower):
                if cell_lower in ("date", "month") and date_col is None:
                    date_col = j
                    continue
                if cell_lower == "oy" and oy_col is None:
                    oy_col = j
                    continue
                # Check against Technical Model column patterns
                for pattern, field_name in TECH_MODEL_COLUMNS:
                    if pattern in cell_lower:
                        if field_name not in col_map.values():
                            col_map[j] = field_name
                        elif field_name == "total_metered_kwh":
                            # Additional metered energy column → sub-meter
                            # Store as _sub_meter:<header> for per-meter extraction
                            col_map[j] = f"_sub_meter:{cells[j]}"
                        break

            # Need at least 3 mapped columns plus date + oy
            if date_col is not None and oy_col is not None and len(col_map) >= 3:
                col_map[date_col] = "_date"
                col_map[oy_col] = "_oy"

                # Also detect meter opening/closing columns by position
                # They are adjacent to invoiced energy columns
                self._detect_meter_reading_cols(cells_lower, col_map)

                logger.debug(
                    f"  Technical Model header at row {i+1}: "
                    f"{len(col_map)} columns mapped"
                )
                return i, col_map

        return None

    def _detect_meter_reading_cols(
        self, header_lower: List[str], col_map: Dict[int, str]
    ) -> None:
        """Detect meter opening/closing columns adjacent to invoiced energy."""
        for j, h in enumerate(header_lower):
            if "opening" in h and "phase" not in h:
                # Determine phase from surrounding context
                # Check if phase-1 invoiced is nearby
                for mapped_j, field in col_map.items():
                    if field == "phase1_invoiced_kwh" and abs(mapped_j - j) <= 2:
                        col_map[j] = "phase1_meter_opening"
                        break
                    elif field == "phase2_invoiced_kwh" and abs(mapped_j - j) <= 2:
                        col_map[j] = "phase2_meter_opening"
                        break
            elif "closing" in h and "phase" not in h:
                for mapped_j, field in col_map.items():
                    if field == "phase1_invoiced_kwh" and abs(mapped_j - j) <= 2:
                        col_map[j] = "phase1_meter_closing"
                        break
                    elif field == "phase2_invoiced_kwh" and abs(mapped_j - j) <= 2:
                        col_map[j] = "phase2_meter_closing"
                        break
            # Phase-specific opening/closing
            if "phase" in h and "1" in h and "open" in h:
                col_map.setdefault(j, "phase1_meter_opening")
            elif "phase" in h and "1" in h and "clos" in h:
                col_map.setdefault(j, "phase1_meter_closing")
            elif "phase" in h and "2" in h and "open" in h:
                col_map.setdefault(j, "phase2_meter_opening")
            elif "phase" in h and "2" in h and "clos" in h:
                col_map.setdefault(j, "phase2_meter_closing")

    def _parse_technical_model(
        self, rows: List[List[Any]], sheet_name: str
    ) -> List[TechnicalModelRow]:
        """Parse the Technical Model section of a sheet."""
        result = self._find_technical_model_header(rows)
        if not result:
            return []

        header_idx, col_map = result
        tech_rows: List[TechnicalModelRow] = []

        for i in range(header_idx + 1, len(rows)):
            row = rows[i]
            if not row or all(v is None for v in row):
                continue

            # Extract date
            date_col = None
            oy_col = None
            for j, field in col_map.items():
                if field == "_date":
                    date_col = j
                elif field == "_oy":
                    oy_col = j

            if date_col is None or oy_col is None:
                continue
            if date_col >= len(row) or oy_col >= len(row):
                continue

            month_date = _parse_date(row[date_col])
            if not month_date:
                continue

            oy_val = _safe_float_allow_zero(row[oy_col])
            if oy_val is None:
                continue
            operating_year = int(oy_val)

            # Extract all mapped fields
            values: Dict[str, Any] = {}
            sub_meters: Dict[str, Any] = {}
            for j, field in col_map.items():
                if field == "_date" or field == "_oy":
                    continue
                if field.startswith("_sub_meter:"):
                    # Named sub-meter column → store in sub_meter_kwh dict
                    if j < len(row):
                        header_name = field[len("_sub_meter:"):]
                        sub_meters[header_name] = _safe_float(row[j])
                    continue
                if j >= len(row):
                    continue

                if field == "comments":
                    val = row[j]
                    values[field] = str(val).strip() if val else None
                elif field in ("forecast_pr", "forecast_pr_poa", "actual_pr", "actual_availability_pct"):
                    val = _safe_float(row[j])
                    if val is not None:
                        values[field] = val if val <= 1.0 else val / 100.0
                elif field in ("energy_comparison", "irr_comparison", "pr_comparison"):
                    val = _safe_float(row[j])
                    if val is not None:
                        values[field] = val if val <= 2.0 else val / 100.0
                elif field in ("phase1_meter_opening", "phase1_meter_closing",
                               "phase2_meter_opening", "phase2_meter_closing"):
                    values[field] = _safe_float_allow_zero(row[j])
                else:
                    values[field] = _safe_float(row[j])

            if sub_meters:
                values["sub_meter_kwh"] = sub_meters

            # Need at least one data value beyond date/oy
            data_values = [v for k, v in values.items() if k != "sub_meter_kwh"]
            has_sub_data = any(v is not None for v in sub_meters.values())
            if not any(v is not None for v in data_values) and not has_sub_data:
                continue

            tech_rows.append(TechnicalModelRow(
                month=month_date,
                operating_year=operating_year,
                **values,
            ))

        return tech_rows

    # ─── Site Parameters Extraction ───────────────────────────────────────

    def _extract_site_parameters(
        self, rows: List[List[Any]], sheet_name: str
    ) -> Dict[str, Any]:
        """Extract site parameters from the header rows (typically rows 2-6)."""
        params: Dict[str, Any] = {}
        phases: Dict[str, Dict[str, Any]] = {}

        for row in rows[:10]:
            if not row:
                continue
            for j, val in enumerate(row):
                if val is None:
                    continue
                label = str(val).strip().lower()

                # Check for phase labels
                phase_key = None
                if "phase 1" in label or "phase1" in label:
                    phase_key = "phase1"
                elif "phase 2" in label or "phase2" in label:
                    phase_key = "phase2"
                elif "combined" in label or "total" in label:
                    phase_key = "combined"

                # Check for parameter labels
                for pattern, param_name in SITE_PARAM_LABELS.items():
                    if pattern in label:
                        # Value is typically the next cell
                        num_val = None
                        if j + 1 < len(row):
                            num_val = _safe_float(row[j + 1])
                        if num_val is None and j + 2 < len(row):
                            num_val = _safe_float(row[j + 2])

                        if num_val is not None:
                            if phase_key:
                                phases.setdefault(phase_key, {})[param_name] = num_val
                            else:
                                params[param_name] = num_val
                        break

        if phases:
            params["phases"] = phases

            # Derive combined capacity if not explicitly present
            if "capacity_kwp" not in params:
                p1_cap = phases.get("phase1", {}).get("capacity_kwp")
                p2_cap = phases.get("phase2", {}).get("capacity_kwp")
                if p1_cap and p2_cap:
                    params["capacity_kwp"] = p1_cap + p2_cap
                elif phases.get("combined", {}).get("capacity_kwp"):
                    params["capacity_kwp"] = phases["combined"]["capacity_kwp"]

        if params:
            logger.debug(f"  {sheet_name}: site params: {params}")

        return params

    # ─── Summary-Performance Tab Parsing ────────────────────────────────

    # Block header keywords for the Summary-Performance tab
    SUMMARY_BLOCK_HEADERS = [
        ("ACTUAL INVOICED ENERGY", "actual_invoiced_energy_kwh"),
        ("EXPECTED OUTPUT", "expected_output_kwh"),
        ("VARIANCE", None),  # Skip variance blocks — set current_block_field=None
        ("ACTUAL IRRADIANCE", "actual_irradiance_wm2"),
        ("EXPECTED IRRADIANCE", "expected_irradiance_wm2"),
        ("EXPECTED POA", "expected_poa_irradiance_wm2"),
        ("PLANT AVAILABILITY", "plant_availability_pct"),
        ("EXPECTED PR POA", "expected_pr_poa_pct"),
        ("EXPECTED PR", "expected_pr_pct"),
    ]

    # Keywords that indicate a summary/total row — NOT a block header
    SUMMARY_SKIP_KEYWORDS = ("TOTAL", "AGREGATED", "WEIGHTED AVERAGE")

    def parse_summary_performance(
        self, project_filter: Optional[str] = None,
    ) -> Dict[str, List[SummaryPerformanceRow]]:
        """
        Parse the 'Summary - Performance' tab.

        Returns dict keyed by sage_id -> list of SummaryPerformanceRow.
        The tab has a block layout: repeated blocks of ~30 project rows,
        each block for a different metric.
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for .xlsx file support")

        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Workbook not found: {self.file_path}")

        wb = openpyxl.load_workbook(self.file_path, data_only=True, read_only=True)

        # Find the Summary-Performance sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "summary" in name.lower() and "performance" in name.lower():
                sheet_name = name
                break

        if not sheet_name:
            wb.close()
            logger.warning("No 'Summary - Performance' tab found")
            return {}

        logger.info(f"Parsing Summary-Performance tab: '{sheet_name}'")
        rows: List[List[Any]] = []
        for row in wb[sheet_name].iter_rows(values_only=True):
            rows.append(list(row))
        wb.close()

        if len(rows) < 5:
            return {}

        # Detect month columns from row near the top (rows 1-5)
        month_cols = self._detect_month_columns(rows)
        if not month_cols:
            logger.warning("Could not detect month columns in Summary-Performance tab")
            return {}

        logger.info(f"  Detected {len(month_cols)} month columns")

        # Detect blocks and parse project rows within each block
        # result: sage_id -> {month -> partial SummaryPerformanceRow fields}
        merged: Dict[str, Dict[date, Dict[str, Any]]] = {}

        current_block_field: Optional[str] = None

        for i, row in enumerate(rows):
            if not row:
                continue

            # Check if this row is a block header
            row_text = " ".join(
                str(c).strip() for c in row[:6] if c is not None
            ).upper()

            # Skip summary/total rows — they contain block keywords
            # but are NOT actual block headers (e.g. "TOTAL AGREGATED
            # ENERGY SALES - ACTUAL INVOICED ENERGY")
            if any(kw in row_text for kw in self.SUMMARY_SKIP_KEYWORDS):
                continue

            block_match = None
            matched_header = False
            for header_keyword, field_name in self.SUMMARY_BLOCK_HEADERS:
                if header_keyword in row_text:
                    block_match = field_name  # None for VARIANCE → skips data rows
                    matched_header = True
                    break

            if matched_header:
                current_block_field = block_match
                logger.debug(f"  Row {i+1}: block header -> {current_block_field}")
                continue

            if current_block_field is None:
                continue

            # Try to read this row as a project data row
            # sage_id is typically in col D or E (index 3 or 4)
            sage_id = self._extract_sage_id_from_summary_row(row)
            if not sage_id:
                continue

            if project_filter and sage_id != project_filter:
                continue

            # Read monthly values
            for col_idx, month_date in month_cols.items():
                if col_idx >= len(row):
                    continue
                val = _safe_float(row[col_idx])
                if val is None:
                    continue

                # Normalize percentage fields
                if current_block_field in ("plant_availability_pct", "expected_pr_pct", "expected_pr_poa_pct"):
                    if val > 1.0:
                        val = val / 100.0

                merged.setdefault(sage_id, {}).setdefault(month_date, {})[
                    current_block_field
                ] = val

        # Convert merged dict to SummaryPerformanceRow list
        result: Dict[str, List[SummaryPerformanceRow]] = {}
        for sage_id, months in merged.items():
            rows_out = []
            for month_date in sorted(months.keys()):
                fields = months[month_date]
                rows_out.append(SummaryPerformanceRow(
                    sage_id=sage_id,
                    month=month_date,
                    **fields,
                ))
            result[sage_id] = rows_out

        total_rows = sum(len(v) for v in result.values())
        logger.info(
            f"  Parsed {total_rows} summary rows for "
            f"{len(result)} project(s)"
        )
        return result

    def _detect_month_columns(
        self, rows: List[List[Any]]
    ) -> Dict[int, date]:
        """
        Detect which columns contain monthly time-series data.

        Scans the first ~10 rows for date headers (datetime objects or
        "Jan-20", "Feb-20" style strings) and returns {col_index: date}.
        """
        for i, row in enumerate(rows[:10]):
            if not row:
                continue
            month_map: Dict[int, date] = {}
            for j, val in enumerate(row):
                if j < 4:  # Skip project info columns (A-D)
                    continue
                d = _parse_date(val)
                if d:
                    month_map[j] = d

            # If we found a reasonable number of month columns, use this row
            if len(month_map) >= 6:
                return month_map

        return {}

    def _extract_sage_id_from_summary_row(self, row: List[Any]) -> Optional[str]:
        """
        Extract sage_id from a project row in the Summary tab.

        Checks cols D and E (index 3, 4) for known sage_ids or tab name mappings.
        """
        for col_idx in (4, 3, 5):  # Prefer col E, then D, then F
            if col_idx >= len(row) or row[col_idx] is None:
                continue
            val = str(row[col_idx]).strip()
            if not val or val in ("-", "N/A", "n/a"):
                continue

            # Direct sage_id match
            if val in TAB_NAME_TO_SAGE_ID.values():
                return val

            # Tab name -> sage_id
            mapped = TAB_NAME_TO_SAGE_ID.get(val)
            if mapped:
                return mapped

        return None

    # ─── Project Waterfall Tab Parsing ────────────────────────────────────

    def parse_project_waterfall(
        self, project_filter: Optional[str] = None,
    ) -> List[WaterfallRow]:
        """
        Parse the 'Project Waterfall' tab.

        Returns list of WaterfallRow, one per project.
        Layout: Col A = sage_id, Col B = kWp, Col C = expected energy,
        Col D = actual energy, Col J = $/kWh tariff.
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for .xlsx file support")

        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Workbook not found: {self.file_path}")

        wb = openpyxl.load_workbook(self.file_path, data_only=True, read_only=True)

        # Find the Project Waterfall sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "waterfall" in name.lower():
                sheet_name = name
                break

        if not sheet_name:
            wb.close()
            logger.warning("No 'Project Waterfall' tab found")
            return []

        logger.info(f"Parsing Project Waterfall tab: '{sheet_name}'")
        rows: List[List[Any]] = []
        for row in wb[sheet_name].iter_rows(values_only=True):
            rows.append(list(row))
        wb.close()

        if len(rows) < 2:
            return []

        # Find header row
        header_idx = None
        for i, row in enumerate(rows[:10]):
            if not row:
                continue
            row_text = " ".join(str(c).strip().lower() for c in row[:5] if c is not None)
            if any(kw in row_text for kw in ["tab id", "project", "sage", "site"]):
                header_idx = i
                break

        # If no header found, assume row 0 is header
        if header_idx is None:
            header_idx = 0

        result: List[WaterfallRow] = []
        for i in range(header_idx + 1, len(rows)):
            row = rows[i]
            if not row or all(v is None for v in row):
                continue

            # Col A: sage_id / Tab ID
            sage_id_raw = row[0] if len(row) > 0 else None
            if sage_id_raw is None:
                continue
            sage_id_str = str(sage_id_raw).strip()
            if not sage_id_str or sage_id_str.lower() in ("total", "grand total", ""):
                continue

            # Skip garbage rows
            if sage_id_str.lower() in ("tab id", "#ref!", "total", "grand total", ""):
                continue

            # Resolve via mapping
            sage_id = TAB_NAME_TO_SAGE_ID.get(sage_id_str, sage_id_str)

            if project_filter and sage_id != project_filter:
                continue

            # Col B: kWp installed capacity
            capacity = _safe_float(row[1]) if len(row) > 1 else None
            # Col C: Expected energy
            expected = _safe_float(row[2]) if len(row) > 2 else None
            # Col D: Actual energy
            actual = _safe_float(row[3]) if len(row) > 3 else None
            # Col J (index 9): $/kWh tariff
            tariff = _safe_float(row[9]) if len(row) > 9 else None

            result.append(WaterfallRow(
                sage_id=sage_id,
                installed_capacity_kwp=capacity,
                expected_energy_kwh=expected,
                actual_energy_kwh=actual,
                tariff_rate_per_kwh=tariff,
            ))

        logger.info(f"  Parsed {len(result)} waterfall rows")
        return result

    # ─── Keyword-Based Parsing (fallback) ─────────────────────────────────

    def _parse_sheet(
        self, ws: Any, sheet_name: str
    ) -> List[PlantPerformanceMonthly]:
        """Parse a single sheet for monthly performance data."""
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        return self._parse_sheet_from_rows(rows, sheet_name)

    def _parse_sheet_from_rows(
        self, rows: List[List[Any]], sheet_name: str
    ) -> List[PlantPerformanceMonthly]:
        """Parse rows using keyword-based column detection."""
        if len(rows) < 2:
            return []

        # Find header row (row with month names or date columns)
        header_idx = None
        col_types: Dict[int, str] = {}

        for i, row in enumerate(rows[:20]):
            if not row:
                continue
            type_matches = 0
            for j, val in enumerate(row):
                if val is None:
                    continue
                ct = _detect_column_type(str(val))
                if ct:
                    col_types[j] = ct
                    type_matches += 1
            if type_matches >= 2:
                header_idx = i
                break

        if header_idx is None:
            logger.debug(f"  {sheet_name}: No recognizable header found")
            return []

        # Find date/month column
        date_col = None
        for j, val in enumerate(rows[header_idx]):
            if val is None:
                continue
            v = str(val).lower()
            if any(kw in v for kw in ["month", "date", "period"]):
                date_col = j
                break

        # Parse data rows
        monthly: List[PlantPerformanceMonthly] = []
        for i in range(header_idx + 1, len(rows)):
            row = rows[i]
            if not row or all(v is None for v in row):
                continue

            month_date = None
            if date_col is not None and date_col < len(row):
                month_date = _parse_date(row[date_col])

            if not month_date:
                continue

            # Extract values by column type
            metered = None
            available = None
            ghi = None
            poa = None
            pr = None
            avail = None

            for col_idx, col_type in col_types.items():
                if col_idx >= len(row):
                    continue
                val = _safe_float(row[col_idx])
                if val is None:
                    continue
                if col_type == "metered_energy":
                    metered = val
                elif col_type == "available_energy":
                    available = val
                elif col_type == "ghi":
                    ghi = val
                elif col_type == "poa":
                    poa = val
                elif col_type == "pr":
                    pr = val if val <= 1.0 else val / 100.0
                elif col_type == "availability":
                    avail = val if val <= 1.0 else val / 100.0

            if metered is not None or available is not None:
                monthly.append(PlantPerformanceMonthly(
                    month=month_date,
                    metered_energy_kwh=metered,
                    available_energy_kwh=available,
                    ghi_kwh_m2=ghi,
                    poa_kwh_m2=poa,
                    performance_ratio_pct=pr,
                    availability_pct=avail,
                ))

        logger.debug(f"  {sheet_name}: {len(monthly)} monthly records")
        return monthly
