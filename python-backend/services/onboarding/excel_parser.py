"""
Label-anchored Excel parser for onboarding templates.

Scans for known section header labels instead of relying on hardcoded
row/column indices, making it resilient to row insertions across
template versions.
"""

import logging
import re
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from models.onboarding import (
    AssetData,
    ContactData,
    ExcelOnboardingData,
    ForecastMonthData,
    MeterData,
)
from services.onboarding.normalizer import (
    normalize_boolean,
    normalize_currency,
    normalize_energy_sale_type,
    normalize_escalation_type,
    normalize_metering_type,
    normalize_percentage,
    normalize_tariff_structure,
)

logger = logging.getLogger(__name__)


# =============================================================================
# LABEL → FIELD MAPPINGS
# =============================================================================

# Project information section — label text (case-insensitive) → field name
PROJECT_INFO_LABELS = {
    "country code": "external_project_id",
    "project name": "project_name",
    "country": "country",
    "customer": "customer_name",
    "sage id": "sage_id",
    "cod date": "cod_date",
    "cod": "cod_date",
    "commercial operation date": "cod_date",
    "installed dc capacity": "installed_dc_capacity_kwp",
    "dc capacity": "installed_dc_capacity_kwp",
    "installed ac capacity": "installed_ac_capacity_kw",
    "ac capacity": "installed_ac_capacity_kw",
    "google maps": "installation_location_url",
    "installation location": "installation_location_url",
    "location url": "installation_location_url",
}

CUSTOMER_INFO_LABELS = {
    "registered name": "registered_name",
    "registration number": "registration_number",
    "tax pin": "tax_pin",
    "tin": "tax_pin",
    "registered address": "registered_address",
    "customer email": "customer_email",
    "email": "customer_email",
    "customer country": "customer_country",
}

CONTRACT_INFO_LABELS = {
    "contract name": "contract_name",
    "ppa name": "contract_name",
    "contract type": "contract_type_code",
    "contract term": "contract_term_years",
    "term years": "contract_term_years",
    "effective date": "effective_date",
    "contract start date": "effective_date",
    "end date": "end_date",
    "expiry date": "end_date",
    "interconnection voltage": "interconnection_voltage_kv",
    "voltage": "interconnection_voltage_kv",
    "payment security": "payment_security_required",
    "security details": "payment_security_details",
    "fx rate source": "agreed_fx_rate_source",
    "agreed exchange rate": "agreed_fx_rate_source",
}

TARIFF_INFO_LABELS = {
    "tariff structure": "tariff_structure",
    "tariff type": "tariff_structure",
    "energy sale type": "energy_sale_type",
    "escalation type": "escalation_type",
    "escalation": "escalation_type",
    "price adjustment type": "escalation_type",
    "billing currency": "billing_currency",
    "currency": "billing_currency",
    "market ref currency": "market_ref_currency",
    "reference currency": "market_ref_currency",
    "base rate": "base_rate",
    "tariff rate": "base_rate",
    "solar tarrif per kwh": "base_rate",
    "solar tariff per kwh": "base_rate",
    "unit": "unit",
    "discount": "discount_pct",
    "solar discount": "discount_pct",
    "floor rate": "floor_rate",
    "floor price": "floor_rate",
    "floor tarrif": "floor_rate",
    "floor tariff": "floor_rate",
    "ceiling rate": "ceiling_rate",
    "ceiling price": "ceiling_rate",
    "ceiling tarrif": "ceiling_rate",
    "ceiling tariff": "ceiling_rate",
    "cap rate": "ceiling_rate",
    "escalation value": "escalation_value",
    "escalation rate": "escalation_value",
    "price adjustment value": "escalation_value",
    "grp method": "grp_method",
    "grid reference price": "grp_method",
    "payment terms": "payment_terms",
}


class ExcelParser:
    """Label-anchored Excel parser for AM Onboarding Template."""

    def parse(self, file_bytes: bytes, filename: str = "template.xlsx") -> ExcelOnboardingData:
        """
        Parse an Excel onboarding template.

        Args:
            file_bytes: Raw Excel file content.
            filename: Original filename (for logging).

        Returns:
            ExcelOnboardingData with all extracted fields.
        """
        logger.info(f"Parsing Excel onboarding template: {filename}")
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)

        data = ExcelOnboardingData()
        sheets = self._find_data_sheets(wb)

        logger.info(
            f"Sheets found: pricing={sheets['pricing'] is not None}, "
            f"technical={sheets['technical'] is not None}, "
            f"yield={sheets['yield'] is not None}"
        )

        # Sheet 1: Pricing & Payment Info (project, customer, contract, tariff, contacts)
        if sheets["pricing"]:
            idx = self._build_cell_index(sheets["pricing"])
            self._extract_labeled_fields(idx, PROJECT_INFO_LABELS, data)
            self._extract_labeled_fields(idx, CUSTOMER_INFO_LABELS, data)
            self._extract_labeled_fields(idx, CONTRACT_INFO_LABELS, data)
            self._extract_labeled_fields(idx, TARIFF_INFO_LABELS, data)
            self._normalize_fields(data)
            data.contacts = self._extract_contacts(sheets["pricing"], idx)
        else:
            logger.warning("No pricing sheet found — returning empty data")
            wb.close()
            return data

        # Sheet 2: Technical Information (meters, assets, capacity, voltage)
        if sheets["technical"]:
            tech_idx = self._build_cell_index(sheets["technical"])
            # Extract capacity fields from technical sheet if not already found
            if not data.installed_dc_capacity_kwp:
                self._extract_labeled_fields(tech_idx, PROJECT_INFO_LABELS, data)
            # Extract contract fields (interconnection voltage) from technical sheet
            if not data.interconnection_voltage_kv:
                self._extract_labeled_fields(tech_idx, CONTRACT_INFO_LABELS, data)
            data.meters = self._extract_meters(sheets["technical"], tech_idx)
            data.assets = self._extract_assets(sheets["technical"], tech_idx)

        # Sheet 3: Yield Report (forecasts)
        if sheets["yield"]:
            yield_idx = self._build_cell_index(sheets["yield"])
            data.forecasts = self._extract_forecasts(sheets["yield"], yield_idx)

        wb.close()

        logger.info(
            f"Excel parsing complete: project={data.project_name}, "
            f"contacts={len(data.contacts)}, meters={len(data.meters)}, "
            f"assets={len(data.assets)}, forecasts={len(data.forecasts)}"
        )
        return data

    # =========================================================================
    # SHEET DISCOVERY
    # =========================================================================

    def _find_data_sheets(self, wb) -> Dict[str, Optional[Worksheet]]:
        """Map logical roles to worksheets."""
        sheets = wb.sheetnames
        result: Dict[str, Optional[Worksheet]] = {"pricing": None, "technical": None, "yield": None}

        if not sheets:
            return result

        for sname in sheets:
            lower = sname.lower()
            if "pricing" in lower or "payment" in lower:
                result["pricing"] = wb[sname]
            elif "technical" in lower and "extract" not in lower:
                result["technical"] = wb[sname]
            elif "yield" in lower or "pvsyst" in lower:
                result["yield"] = wb[sname]

        # Fallback: if no pricing sheet found, try legacy single-sheet heuristics
        if result["pricing"] is None:
            for pname in ["onboarding", "project", "data", "template", "input", "main"]:
                for sname in sheets:
                    if pname in sname.lower():
                        result["pricing"] = wb[sname]
                        break
                if result["pricing"]:
                    break
            if result["pricing"] is None:
                result["pricing"] = wb[sheets[0]]

        return result

    # =========================================================================
    # CELL INDEX
    # =========================================================================

    def _build_cell_index(self, ws: Worksheet) -> dict:
        """
        Build a dict mapping lowercase label text → (row, col, value_cell_value).

        Detects structured templates (Description | Guidance | Data | Comments)
        and uses column C for values instead of "next non-empty cell to the right".
        """
        # First pass: detect if sheet uses a structured columnar layout
        # by looking for a header row with "DESCRIPTION" in col A and "DETAILS" in col C
        structured_layout = False
        data_col = 2  # Default: column C (0-indexed)
        for row in ws.iter_rows(max_row=15):
            cells = [c.value for c in row]
            if len(cells) > 2:
                col_a = str(cells[0]).strip().upper() if cells[0] else ""
                col_c = str(cells[2]).strip().upper() if len(cells) > 2 and cells[2] else ""
                if "DESCRIPTION" in col_a and "DETAILS" in col_c:
                    structured_layout = True
                    break

        if structured_layout:
            logger.debug("Detected structured columnar layout (Description | Guidance | Data)")

        index = {}
        for row in ws.iter_rows():
            cells = [c.value for c in row]
            if not any(cells):
                continue

            for i, cell_val in enumerate(cells):
                if cell_val is None:
                    continue
                text = str(cell_val).strip()
                if not text:
                    continue

                key = text.lower()

                if structured_layout and i == 0:
                    # In structured layout, col A = label, col C = value
                    value = cells[data_col] if len(cells) > data_col else None
                    index[key] = {
                        "row": row[0].row if hasattr(row[0], 'row') else 0,
                        "col": 1,
                        "label": text,
                        "value": value,
                    }
                elif not structured_layout:
                    # Legacy: value is next non-empty cell to the right
                    value = None
                    for j in range(i + 1, len(cells)):
                        if cells[j] is not None:
                            value = cells[j]
                            break
                    index[key] = {
                        "row": row[0].row if hasattr(row[0], 'row') else 0,
                        "col": i + 1,
                        "label": text,
                        "value": value,
                    }

        return index

    # =========================================================================
    # LABEL-BASED EXTRACTION
    # =========================================================================

    def _extract_labeled_fields(
        self,
        cell_index: dict,
        label_map: dict,
        data: ExcelOnboardingData,
    ) -> None:
        """
        For each label in label_map, find it in cell_index and set
        the corresponding field on data.
        """
        for label_text, field_name in label_map.items():
            match = self._find_label_match(cell_index, label_text)
            if match is None:
                continue

            raw_value = match["value"]
            if raw_value is None:
                continue

            # Skip placeholder/guidance values
            if isinstance(raw_value, str) and self._is_placeholder(raw_value):
                continue

            converted = self._convert_value(field_name, raw_value)
            if converted is not None and hasattr(data, field_name):
                setattr(data, field_name, converted)

    def _find_label_match(self, cell_index: dict, label: str) -> Optional[dict]:
        """Find the best match for a label in the cell index."""
        label_lower = label.lower()

        # Exact match
        if label_lower in cell_index:
            return cell_index[label_lower]

        # Substring match — find keys containing the label
        for key, entry in cell_index.items():
            if label_lower in key:
                return entry

        return None

    def _convert_value(self, field_name: str, raw_value: Any) -> Any:
        """Convert a raw cell value to the appropriate Python type."""
        if raw_value is None:
            return None

        # Date fields
        if field_name in ("cod_date", "effective_date", "end_date"):
            return self._to_date(raw_value)

        # Numeric fields
        if field_name in (
            "installed_dc_capacity_kwp", "installed_ac_capacity_kw",
            "interconnection_voltage_kv", "base_rate", "floor_rate",
            "ceiling_rate", "escalation_value",
        ):
            return self._to_float(raw_value)

        # Integer fields
        if field_name in ("contract_term_years",):
            return self._to_int(raw_value)

        # Percentage fields
        if field_name in ("discount_pct",):
            return normalize_percentage(raw_value)

        # Boolean fields
        if field_name in ("payment_security_required",):
            return normalize_boolean(raw_value)

        # Default: string
        return str(raw_value).strip() if raw_value else None

    # =========================================================================
    # NORMALIZATION (post-extraction)
    # =========================================================================

    def _normalize_fields(self, data: ExcelOnboardingData) -> None:
        """Apply code normalization to free-text fields."""
        data.tariff_structure = normalize_tariff_structure(data.tariff_structure) or data.tariff_structure
        data.energy_sale_type = normalize_energy_sale_type(data.energy_sale_type) or data.energy_sale_type
        data.escalation_type = normalize_escalation_type(data.escalation_type) or data.escalation_type
        data.billing_currency = normalize_currency(data.billing_currency) or data.billing_currency
        data.market_ref_currency = normalize_currency(data.market_ref_currency) or data.market_ref_currency

    # =========================================================================
    # TABLE SECTIONS
    # =========================================================================

    def _extract_contacts(self, ws: Worksheet, cell_index: dict) -> List[ContactData]:
        """Extract contacts from the Customer Contacts section."""
        contacts = []
        header_row = self._find_section_header_row(cell_index, "customer contacts")
        if header_row is None:
            header_row = self._find_section_header_row(cell_index, "contacts")
        if header_row is None:
            return contacts

        for row in ws.iter_rows(min_row=header_row + 1):
            cells = [c.value for c in row]
            # Skip empty rows
            if not any(cells):
                continue

            # Detect end of section (next section header or blank gap)
            first_val = str(cells[0]).strip().lower() if cells[0] else ""
            if first_val and any(kw in first_val for kw in ("document", "asset", "equipment", "meter", "forecast")):
                break

            # Layout: A=Role, B=Invoice Flag, C=Full Name, D=Email, E=Phone
            role = str(cells[0]).strip() if cells[0] else None
            invoice_flag = normalize_boolean(cells[1]) if len(cells) > 1 else False
            full_name = str(cells[2]).strip() if len(cells) > 2 and cells[2] else None
            email = str(cells[3]).strip() if len(cells) > 3 and cells[3] else None
            phone = str(cells[4]).strip() if len(cells) > 4 and cells[4] else None

            # Skip header rows and empty data
            if not full_name and not email:
                continue
            if full_name and full_name.lower() in ("full name", "name", "contact name"):
                continue

            contacts.append(ContactData(
                role=role,
                include_in_invoice=invoice_flag or False,
                full_name=full_name,
                email=email,
                phone=phone,
            ))

        logger.info(f"Extracted {len(contacts)} contacts")
        return contacts

    def _extract_meters(self, ws: Worksheet, cell_index: dict) -> List[MeterData]:
        """Extract meters by splitting comma-separated serial numbers."""
        meters = []

        # Look for meter serial numbers field
        serial_match = self._find_label_match(cell_index, "meter serial")
        if serial_match is None:
            serial_match = self._find_label_match(cell_index, "serial number")
        if serial_match is None:
            return meters

        raw_serials = serial_match.get("value")
        if not raw_serials:
            return meters

        # Split comma-separated
        serials = [s.strip() for s in str(raw_serials).split(",") if s.strip()]

        # Look for metering type
        metering_match = self._find_label_match(cell_index, "metering type")
        metering_type = None
        if metering_match and metering_match.get("value"):
            metering_type = normalize_metering_type(str(metering_match["value"]))

        for serial in serials:
            meters.append(MeterData(
                serial_number=serial,
                metering_type=metering_type,
            ))

        # Cross-check with "Number of Meters" field
        count_match = self._find_label_match(cell_index, "number of meters")
        if count_match and count_match.get("value"):
            expected = self._to_int(count_match["value"])
            if expected and expected != len(meters):
                logger.warning(
                    f"Meter count mismatch: 'Number of Meters'={expected}, "
                    f"parsed serial count={len(meters)}"
                )

        logger.info(f"Extracted {len(meters)} meters")
        return meters

    def _extract_assets(self, ws: Worksheet, cell_index: dict) -> List[AssetData]:
        """Extract equipment/asset data from the installation section."""
        assets = []
        header_row = self._find_section_header_row(cell_index, "equipment")
        if header_row is None:
            header_row = self._find_section_header_row(cell_index, "installation")
        if header_row is None:
            header_row = self._find_section_header_row(cell_index, "asset")
        if header_row is None:
            return assets

        for row in ws.iter_rows(min_row=header_row + 1):
            cells = [c.value for c in row]
            if not any(cells):
                continue

            first_val = str(cells[0]).strip().lower() if cells[0] else ""
            if first_val and any(kw in first_val for kw in ("document", "contact", "meter", "forecast")):
                break

            asset_type = str(cells[0]).strip() if cells[0] else None
            if not asset_type:
                continue

            # Map common asset type names to codes
            type_code = self._map_asset_type(asset_type)

            assets.append(AssetData(
                asset_type_code=type_code,
                asset_name=str(cells[1]).strip() if len(cells) > 1 and cells[1] else None,
                model=str(cells[2]).strip() if len(cells) > 2 and cells[2] else None,
                serial_code=str(cells[3]).strip() if len(cells) > 3 and cells[3] else None,
                capacity=self._to_float(cells[4]) if len(cells) > 4 else None,
                capacity_unit=str(cells[5]).strip() if len(cells) > 5 and cells[5] else None,
                quantity=self._to_int(cells[6]) or 1 if len(cells) > 6 else 1,
            ))

        logger.info(f"Extracted {len(assets)} assets")
        return assets

    def _extract_forecasts(self, ws: Worksheet, cell_index: dict) -> List[ForecastMonthData]:
        """Extract monthly production forecasts."""
        forecasts = []

        # Try standard section header first
        header_row = self._find_section_header_row(cell_index, "forecast")
        if header_row is None:
            header_row = self._find_section_header_row(cell_index, "production")

        if header_row is not None:
            forecasts = self._extract_forecasts_standard(ws, header_row)

        # Try PVSyst-style yield report if standard extraction failed
        if not forecasts:
            forecasts = self._extract_forecasts_pvsyst(ws, cell_index)

        logger.info(f"Extracted {len(forecasts)} forecast months")
        return forecasts

    def _extract_forecasts_standard(self, ws: Worksheet, header_row: int) -> List[ForecastMonthData]:
        """Extract forecasts from a standard tabular format."""
        forecasts = []
        for row in ws.iter_rows(min_row=header_row + 1):
            cells = [c.value for c in row]
            if not any(cells):
                continue

            first_val = str(cells[0]).strip().lower() if cells[0] else ""
            if first_val and any(kw in first_val for kw in ("document", "contact", "guarantee", "total")):
                break

            month = self._to_date(cells[0])
            energy = self._to_float(cells[1]) if len(cells) > 1 else None
            if month is None or energy is None:
                continue

            forecasts.append(ForecastMonthData(
                forecast_month=month,
                operating_year=self._to_int(cells[2]) if len(cells) > 2 else None,
                forecast_energy_kwh=energy,
                forecast_ghi=self._to_float(cells[3]) if len(cells) > 3 else None,
                forecast_poa=self._to_float(cells[4]) if len(cells) > 4 else None,
                forecast_pr=self._to_float(cells[5]) if len(cells) > 5 else None,
                degradation_factor=self._to_float(cells[6]) if len(cells) > 6 else None,
            ))
        return forecasts

    # Month name abbreviations for PVSyst-style yield reports
    _MONTH_NAMES = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4,
        "may": 5, "jun": 6, "jul": 7, "aug": 8,
        "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }

    def _extract_forecasts_pvsyst(self, ws: Worksheet, cell_index: dict) -> List[ForecastMonthData]:
        """Extract forecasts from PVSyst-style yield report with month names and multi-column layout."""
        forecasts = []

        # Find the header row with column labels (look for "E_Grid" or "GlobHor")
        header_row_idx = None
        energy_col = None
        ghi_col = None
        poa_col = None
        pr_col = None

        for i, row in enumerate(ws.iter_rows(max_row=15)):
            cells = {j: str(c.value).strip().lower() if c.value else "" for j, c in enumerate(row)}
            for j, val in cells.items():
                if val in ("e_grid", "energy output"):
                    header_row_idx = i + 1
                    energy_col = j
                elif val in ("globhor", "ghi irr", "ghi"):
                    ghi_col = j
                elif val in ("globinc", "poa irr", "poa"):
                    poa_col = j
                elif val == "pr" or val == "pr*":
                    pr_col = j

        if header_row_idx is None or energy_col is None:
            return forecasts

        # Use a reference year (doesn't matter much for monthly forecasts)
        ref_year = 2025

        for row in ws.iter_rows(min_row=header_row_idx + 1):
            cells = [c.value for c in row]
            if not cells or cells[0] is None:
                continue

            first_val = str(cells[0]).strip().lower()

            # Stop at "year" or "total" summary row
            if first_val in ("year", "total", "annual"):
                break

            # Match month names
            month_num = self._MONTH_NAMES.get(first_val[:3])
            if month_num is None:
                continue

            energy = self._to_float(cells[energy_col]) if energy_col < len(cells) else None
            if energy is None:
                # Try E_Grid column (col 7 in typical PVSyst layout)
                for try_col in [7, 6, 16]:
                    if try_col < len(cells):
                        energy = self._to_float(cells[try_col])
                        if energy and energy > 1000:  # Sanity check: monthly energy should be > 1000 kWh
                            break

            if energy is None:
                continue

            forecast_month = date(ref_year, month_num, 1)
            ghi = self._to_float(cells[ghi_col]) if ghi_col and ghi_col < len(cells) else None
            poa = self._to_float(cells[poa_col]) if poa_col and poa_col < len(cells) else None
            pr = self._to_float(cells[pr_col]) if pr_col and pr_col < len(cells) else None

            forecasts.append(ForecastMonthData(
                forecast_month=forecast_month,
                operating_year=1,
                forecast_energy_kwh=energy,
                forecast_ghi=ghi,
                forecast_poa=poa,
                forecast_pr=pr,
            ))

        return forecasts

    # =========================================================================
    # HELPERS
    # =========================================================================

    @staticmethod
    def _is_placeholder(value: str) -> bool:
        """Return True if the value looks like template guidance or placeholder text."""
        lower = value.strip().lower()
        placeholders = [
            "select from dropdown",
            "per project operations",
            "am to input",
            "to be completed",
            "insert #",
            "enter each product",
            "please indicate",
            "please include",
            "if applicable",
            "y/n",
            "dd-month-yy",
            "guidence",
            "guidance",
        ]
        return any(p in lower for p in placeholders)

    def _find_section_header_row(self, cell_index: dict, keyword: str) -> Optional[int]:
        """Find the row number of a section header containing keyword."""
        keyword_lower = keyword.lower()
        for key, entry in cell_index.items():
            if keyword_lower in key:
                return entry["row"]
        return None

    @staticmethod
    def _map_asset_type(name: str) -> str:
        """Map common asset type names to database codes."""
        name_lower = name.lower()
        mappings = {
            "panel": "SOLAR_PANEL",
            "module": "SOLAR_PANEL",
            "solar panel": "SOLAR_PANEL",
            "inverter": "INVERTER",
            "string inverter": "INVERTER",
            "central inverter": "INVERTER",
            "transformer": "TRANSFORMER",
            "meter": "METER",
            "battery": "BATTERY",
            "bess": "BATTERY",
            "tracker": "TRACKER",
            "mounting": "MOUNTING_STRUCTURE",
            "structure": "MOUNTING_STRUCTURE",
            "combiner": "COMBINER_BOX",
            "combiner box": "COMBINER_BOX",
        }
        for key, code in mappings.items():
            if key in name_lower:
                return code
        return name.upper().replace(" ", "_")

    @staticmethod
    def _to_date(value: Any) -> Optional[date]:
        """Convert various date representations to date."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        try:
            # Try common formats
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(str(value).strip(), fmt).date()
                except ValueError:
                    continue
        except Exception:
            pass
        return None

    @staticmethod
    def _to_float(value: Any) -> Optional[float]:
        """Convert to float, handling commas and units."""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        try:
            cleaned = str(value).replace(",", "").replace(" ", "").strip()
            # Remove common units
            cleaned = re.sub(r'(kwp|kw|kwh|mwh|mw|kva|%|usd|ghs)$', '', cleaned, flags=re.IGNORECASE)
            return float(cleaned)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _to_int(value: Any) -> Optional[int]:
        """Convert to integer."""
        if value is None:
            return None
        if isinstance(value, int):
            return value
        try:
            return int(float(str(value).replace(",", "").strip()))
        except (ValueError, TypeError):
            return None
