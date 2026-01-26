"""
Excel XLSX output formatter.

Creates Excel workbooks with multiple sheets for headers and line items.
"""

import io
import logging
from datetime import datetime, date
from decimal import Decimal
from typing import Dict, Any, List, Optional

from models.reports import ExtractedData, FileFormat, InvoiceReportType
from .base import BaseFormatter

logger = logging.getLogger(__name__)

# Lazy import openpyxl to avoid import errors if not installed
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed - XLSX formatting unavailable")


class XLSXFormatter(BaseFormatter):
    """
    Formatter that outputs report data as Excel XLSX.

    Supports:
    - Multiple worksheets (Summary, Invoices, Line Items)
    - Header row formatting (bold, background color)
    - Currency formatting
    - Auto-column width
    """

    def __init__(self):
        """Initialize formatter and check dependencies."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for XLSX formatting. "
                "Install with: pip install openpyxl"
            )

    def get_file_format(self) -> FileFormat:
        """Return XLSX file format."""
        return FileFormat.XLSX

    def get_content_type(self) -> str:
        """Return Excel MIME type."""
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def get_file_extension(self) -> str:
        """Return xlsx extension."""
        return "xlsx"

    def format(
        self,
        extracted_data: ExtractedData,
        template_config: Dict[str, Any]
    ) -> bytes:
        """
        Format extracted data as Excel XLSX.

        Args:
            extracted_data: The data to format
            template_config: Configuration options:
                - include_summary (bool): Include summary sheet (default: True)
                - include_line_items (bool): Include line items sheet (default: True)
                - freeze_headers (bool): Freeze header row (default: True)

        Returns:
            XLSX file as bytes

        Raises:
            ValueError: If formatting fails
        """
        logger.info(
            f"Formatting {extracted_data.report_type.value} data as XLSX"
        )

        # Get config options
        include_summary = template_config.get('include_summary', True)
        include_line_items = template_config.get('include_line_items', True)
        freeze_headers = template_config.get('freeze_headers', True)

        try:
            wb = Workbook()

            # Remove default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)

            # Create Summary sheet
            if include_summary:
                self._create_summary_sheet(wb, extracted_data)

            # Create Invoices sheet
            self._create_invoices_sheet(wb, extracted_data, freeze_headers)

            # Create Line Items sheet
            if include_line_items and extracted_data.line_items:
                self._create_line_items_sheet(wb, extracted_data, freeze_headers)

            # Create Comparison sheet for comparison reports
            if extracted_data.comparison_data:
                self._create_comparison_sheet(wb, extracted_data)

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()

        except Exception as e:
            logger.error(f"XLSX formatting failed: {e}")
            raise ValueError(f"Failed to format data as XLSX: {e}")

    def _create_summary_sheet(
        self,
        wb: 'Workbook',
        extracted_data: ExtractedData
    ) -> None:
        """Create the summary sheet with metadata."""
        ws = wb.create_sheet("Summary")

        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # Write summary data
        summary_data = [
            ("Report Type", extracted_data.report_type.value),
            ("Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
            ("", ""),
        ]

        if extracted_data.billing_period:
            summary_data.extend([
                ("Billing Period", ""),
                ("  Start Date", self._format_value(extracted_data.billing_period.start_date)),
                ("  End Date", self._format_value(extracted_data.billing_period.end_date)),
                ("", ""),
            ])

        summary_data.extend([
            ("Statistics", ""),
            ("  Total Records", extracted_data.metadata.record_count),
            ("  Total Amount", self._format_value(extracted_data.metadata.total_amount)),
            ("  Contracts", ", ".join(extracted_data.metadata.contract_names) or "N/A"),
        ])

        for row_idx, (label, value) in enumerate(summary_data, start=1):
            ws.cell(row=row_idx, column=1, value=label).font = header_font
            ws.cell(row=row_idx, column=2, value=value)

        # Auto-fit columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

    def _create_invoices_sheet(
        self,
        wb: 'Workbook',
        extracted_data: ExtractedData,
        freeze_headers: bool
    ) -> None:
        """Create the invoices/headers sheet."""
        ws = wb.create_sheet("Invoices")

        if not extracted_data.headers:
            ws.cell(row=1, column=1, value="No invoice data available")
            return

        # Get columns based on report type
        columns = self._get_header_columns(extracted_data.report_type)
        column_labels = self._get_column_labels(columns)

        # Write header row
        self._write_header_row(ws, column_labels, 1)

        # Write data rows
        for row_idx, header in enumerate(extracted_data.headers, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = self._format_value(header.get(col_name))
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit columns and freeze header
        self._auto_fit_columns(ws, len(columns))
        if freeze_headers:
            ws.freeze_panes = 'A2'

    def _create_line_items_sheet(
        self,
        wb: 'Workbook',
        extracted_data: ExtractedData,
        freeze_headers: bool
    ) -> None:
        """Create the line items sheet."""
        ws = wb.create_sheet("Line Items")

        if not extracted_data.line_items:
            ws.cell(row=1, column=1, value="No line item data available")
            return

        # Get columns based on report type
        columns = self._get_line_item_columns(extracted_data.report_type)
        column_labels = self._get_column_labels(columns)

        # Write header row
        self._write_header_row(ws, column_labels, 1)

        # Write data rows
        for row_idx, item in enumerate(extracted_data.line_items, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = self._format_value(item.get(col_name))
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit columns and freeze header
        self._auto_fit_columns(ws, len(columns))
        if freeze_headers:
            ws.freeze_panes = 'A2'

    def _create_comparison_sheet(
        self,
        wb: 'Workbook',
        extracted_data: ExtractedData
    ) -> None:
        """Create the variance summary sheet for comparison reports."""
        ws = wb.create_sheet("Variance Summary")

        # Define styles
        header_font = Font(bold=True)
        positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        comparison = extracted_data.comparison_data or {}

        summary_data = [
            ("Variance Analysis Summary", ""),
            ("", ""),
            ("Total Expected", self._format_value(comparison.get('total_expected'))),
            ("Total Received", self._format_value(comparison.get('total_received'))),
            ("Total Variance", self._format_value(comparison.get('total_variance'))),
            ("Variance %", f"{comparison.get('variance_percentage', 0):.2f}%"),
            ("", ""),
            ("Status Breakdown", ""),
            ("  Matched", comparison.get('matched_count', 0)),
            ("  Overbilled", comparison.get('overbilled_count', 0)),
            ("  Underbilled", comparison.get('underbilled_count', 0)),
        ]

        for row_idx, (label, value) in enumerate(summary_data, start=1):
            cell_label = ws.cell(row=row_idx, column=1, value=label)
            cell_value = ws.cell(row=row_idx, column=2, value=value)

            if label and not label.startswith(" "):
                cell_label.font = header_font

            # Highlight variance rows
            if "Overbilled" in str(label):
                cell_value.fill = negative_fill
            elif "Underbilled" in str(label):
                cell_value.fill = positive_fill

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _write_header_row(
        self,
        ws,
        labels: List[str],
        row: int
    ) -> None:
        """Write a formatted header row."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, label in enumerate(labels, start=1):
            cell = ws.cell(row=row, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _auto_fit_columns(self, ws, num_columns: int) -> None:
        """Auto-fit column widths based on content."""
        for col_idx in range(1, num_columns + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _format_value(self, value: Any) -> Any:
        """Format a value for Excel output."""
        if value is None:
            return ""
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, (datetime, date)):
            return value
        return value

    def _get_header_columns(self, report_type: InvoiceReportType) -> List[str]:
        """Get column names for headers based on report type."""
        base = ['id', 'contract_name', 'project_name']

        if report_type == InvoiceReportType.INVOICE_TO_CLIENT:
            return base + ['invoice_number', 'invoice_date', 'due_date', 'total_amount', 'status']
        elif report_type == InvoiceReportType.INVOICE_EXPECTED:
            return base + ['expected_total_amount', 'created_at']
        elif report_type == InvoiceReportType.INVOICE_RECEIVED:
            return base + ['vendor_invoice_number', 'invoice_date', 'due_date', 'total_amount', 'status']
        elif report_type == InvoiceReportType.INVOICE_COMPARISON:
            return ['id', 'contract_name', 'expected_total_amount', 'received_total_amount',
                    'header_variance', 'comparison_status']
        return base + ['total_amount']

    def _get_line_item_columns(self, report_type: InvoiceReportType) -> List[str]:
        """Get column names for line items based on report type."""
        if report_type == InvoiceReportType.INVOICE_TO_CLIENT:
            return ['id', 'invoice_header_id', 'description', 'quantity',
                    'line_unit_price', 'line_total_amount', 'line_item_type']
        elif report_type == InvoiceReportType.INVOICE_EXPECTED:
            return ['id', 'expected_invoice_header_id', 'description',
                    'line_total_amount', 'line_item_type']
        elif report_type == InvoiceReportType.INVOICE_RECEIVED:
            return ['id', 'received_invoice_header_id', 'description',
                    'line_total_amount', 'line_item_type']
        elif report_type == InvoiceReportType.INVOICE_COMPARISON:
            return ['id', 'invoice_comparison_id', 'expected_description',
                    'expected_line_amount', 'received_description',
                    'received_line_amount', 'variance_amount']
        return ['id', 'description', 'line_total_amount']

    def _get_column_labels(self, columns: List[str]) -> List[str]:
        """Convert column names to human-readable labels."""
        label_map = {
            'id': 'ID',
            'contract_name': 'Contract',
            'project_name': 'Project',
            'invoice_number': 'Invoice #',
            'vendor_invoice_number': 'Vendor Invoice #',
            'invoice_date': 'Invoice Date',
            'due_date': 'Due Date',
            'total_amount': 'Total Amount',
            'expected_total_amount': 'Expected Amount',
            'received_total_amount': 'Received Amount',
            'header_variance': 'Variance',
            'comparison_status': 'Status',
            'status': 'Status',
            'created_at': 'Created',
            'received_date': 'Received Date',
            'invoice_header_id': 'Invoice ID',
            'expected_invoice_header_id': 'Expected Invoice ID',
            'received_invoice_header_id': 'Received Invoice ID',
            'invoice_comparison_id': 'Comparison ID',
            'description': 'Description',
            'expected_description': 'Expected Description',
            'received_description': 'Received Description',
            'quantity': 'Quantity',
            'line_unit_price': 'Unit Price',
            'line_total_amount': 'Amount',
            'expected_line_amount': 'Expected Amount',
            'received_line_amount': 'Received Amount',
            'variance_amount': 'Variance',
            'line_item_type': 'Type',
        }
        return [label_map.get(col, col.replace('_', ' ').title()) for col in columns]
